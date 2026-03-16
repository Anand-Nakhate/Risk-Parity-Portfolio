"""
EDGAR N-PORT Holdings to Excel — Exhaustive Extractor (v2)

Usage:
    python edgar_nport_to_excel.py PBAIX
    python edgar_nport_to_excel.py PBAIX --output holdings.xlsx
    python edgar_nport_to_excel.py PBAIX --date 2024-06-30
    python edgar_nport_to_excel.py PBAIX --all-filings
    python edgar_nport_to_excel.py PBAIX --user-agent "YourName your@email.com"

Extracts EVERY field from SEC EDGAR N-PORT filings and writes a comprehensive
Excel workbook. Covers all instrument types: equities, debt, repos, reverse repos,
forwards, futures, swaps (IRS, CDS, CCS), options, swaptions, warrants,
ABS/MBS, loans, structured notes, and short-term investment vehicles.

Requirements: pip install requests openpyxl
"""

import argparse
import json
import sys
import time
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from datetime import datetime, date
from typing import Optional

import openpyxl
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# Constants
# ============================================================================

RATE_LIMIT_DELAY = 0.12  # ~8 req/s, safely under SEC's 10 req/s limit
MAX_RETRIES = 3
RETRY_BACKOFF_BASE = 2.0
RETRY_STATUS_CODES = {429, 500, 502, 503, 504}

DEFAULT_USER_AGENT = "EdgarNportExtractor your.email@example.com"

ASSET_CATEGORY_MAP = {
    "EC": "Equity - Common Stock",
    "EP": "Equity - Preferred Stock",
    "DBT": "Debt",
    "ABS-MBS": "Asset-Backed - Mortgage-Backed Security",
    "ABS-CBDO": "Asset-Backed - Collateralized Bond/Debt Obligation",
    "ABS-O": "Asset-Backed - Other",
    "LON": "Loan",
    "DE": "Derivative - Equity",
    "DFE": "Derivative - Foreign Exchange",
    "DIR": "Derivative - Interest Rate",
    "DCR": "Derivative - Credit",
    "DCO": "Derivative - Commodity",
    "DO": "Derivative - Other",
    "STIV": "Short-Term Investment Vehicle",
    "SN": "Structured Note",
    "RA": "Repurchase Agreement",
    "OTHER": "Other",
}

ISSUER_CATEGORY_MAP = {
    "CORP": "Corporate",
    "NUSS": "Non-US Sovereign",
    "UST": "US Treasury",
    "USGSE": "US Govt Sponsored Entity",
    "USGA": "US Government Agency",
    "SOVN": "Sovereign",
    "MUN": "Municipal",
    "RF": "Registered Fund",
    "FN": "Financial Institution",
    "OTHER": "Other",
}

UNITS_MAP = {
    "PA": "Principal Amount",
    "NS": "Number of Shares",
    "NC": "Number of Contracts",
    "OU": "Other Units",
}

DERIV_CATEGORY_MAP = {
    "FWD": "Forward",
    "FUT": "Future",
    "SWP": "Swap",
    "OPT": "Option",
    "SWN": "Swaption",
    "WAR": "Warrant",
}

CONTRACT_TYPE_ELEMENTS = [
    ("commodityContracts", "Commodity"),
    ("creditContracts", "Credit"),
    ("equityContracts", "Equity"),
    ("foreignExchgContracts", "Foreign Exchange"),
    ("interestRtContracts", "Interest Rate"),
    ("otherContracts", "Other"),
]

INSTRUMENT_CATEGORIES = [
    ("forwardCategory", "Forward"),
    ("futureCategory", "Future"),
    ("optionCategory", "Option"),
    ("swaptionCategory", "Swaption"),
    ("swapCategory", "Swap"),
    ("warrantCategory", "Warrant"),
    ("otherCategory", "Other"),
]

REPO_COLLATERAL_CATEGORY_MAP = {
    "UST": "US Treasuries",
    "USGA": "US Government Agency Debt",
    "CDS": "Corporate Debt Securities",
    "EC": "Equity - Common",
    "EP": "Equity - Preferred",
    "ABS": "Asset-Backed Securities",
    "Other instrument": "Other Instrument",
}


# ============================================================================
# SEC API Layer
# ============================================================================

def sec_get(url: str, headers: dict) -> requests.Response:
    """Rate-limited GET with exponential backoff retry."""
    time.sleep(RATE_LIMIT_DELAY)
    for attempt in range(MAX_RETRIES + 1):
        try:
            resp = requests.get(url, headers=headers, timeout=30)
            if resp.status_code in RETRY_STATUS_CODES and attempt < MAX_RETRIES:
                wait = RETRY_BACKOFF_BASE ** (attempt + 1)
                if resp.status_code == 429:
                    retry_after = resp.headers.get("Retry-After")
                    if retry_after and retry_after.isdigit():
                        wait = max(wait, int(retry_after))
                print(f"  Retry {attempt + 1}/{MAX_RETRIES} after {wait:.0f}s (HTTP {resp.status_code})...", file=sys.stderr)
                time.sleep(wait)
                continue
            resp.raise_for_status()
            return resp
        except requests.exceptions.ConnectionError:
            if attempt < MAX_RETRIES:
                wait = RETRY_BACKOFF_BASE ** (attempt + 1)
                print(f"  Connection error, retry {attempt + 1}/{MAX_RETRIES} after {wait:.0f}s...", file=sys.stderr)
                time.sleep(wait)
                continue
            raise
    resp.raise_for_status()
    return resp


def lookup_ticker(ticker: str, headers: dict) -> dict:
    """Look up CIK, series ID, and class ID for a mutual fund ticker."""
    print(f"Looking up ticker {ticker}...")
    resp = sec_get("https://www.sec.gov/files/company_tickers_mf.json", headers)
    data = resp.json()["data"]
    for row in data:
        if len(row) >= 4 and str(row[3]).upper() == ticker.upper():
            return {
                "cik": str(row[0]),
                "series_id": row[1],
                "class_id": row[2],
                "ticker": row[3],
            }
    raise ValueError(f"Ticker '{ticker}' not found in SEC mutual fund tickers")


def find_nport_filing(cik: str, series_id: str, headers: dict, target_date: Optional[str] = None) -> dict:
    """Find the correct NPORT-P filing for a specific series.

    Uses EFTS full-text search to find the filing containing the exact series ID.
    This is critical for umbrella CIKs (e.g., PIMCO, BlackRock) that file many
    NPORT-P filings under one CIK, each for a different series/fund.
    Falls back to submissions API if EFTS fails.
    """
    # --- Primary approach: EFTS search by series ID ---
    print(f"Searching for NPORT-P filings for series {series_id}...")
    try:
        start_dt = target_date[:4] + "-01-01" if target_date else "2020-01-01"
        end_dt = "2030-12-31"
        efts_url = (
            f"https://efts.sec.gov/LATEST/search-index"
            f"?q=%22{series_id}%22&forms=NPORT-P"
            f"&dateRange=custom&startdt={start_dt}&enddt={end_dt}"
        )
        resp = sec_get(efts_url, headers)
        hits = resp.json().get("hits", {}).get("hits", [])
        if hits:
            hits.sort(key=lambda h: h["_source"].get("file_date", ""), reverse=True)
            if target_date:
                hits.sort(key=lambda h: abs(_date_diff(
                    h["_source"].get("period_ending") or h["_source"].get("file_date", ""),
                    target_date
                )))
            src = hits[0]["_source"]
            cik_from_efts = src["ciks"][0].lstrip("0") if src.get("ciks") else cik
            xsl = src.get("xsl", "")
            primary_doc = f"{xsl}/primary_doc.xml" if xsl else "primary_doc.xml"
            result = {
                "accession": src["adsh"],
                "file_date": src.get("file_date", ""),
                "report_date": src.get("period_ending", ""),
                "primary_document": primary_doc,
                "cik": cik_from_efts,
            }
            print(f"  Found via EFTS: {result['file_date']}, period={result['report_date']}")
            return result
    except Exception as e:
        print(f"  EFTS search failed ({e}), falling back to submissions API...", file=sys.stderr)

    # --- Fallback: submissions API + iterate to find matching series ---
    padded_cik = cik.zfill(10)
    url = f"https://data.sec.gov/submissions/CIK{padded_cik}.json"
    print(f"  Fetching filing history from {url}...")
    resp = sec_get(url, headers)
    data = resp.json()

    nport_filings = _extract_nport_filings(data.get("filings", {}).get("recent", {}), cik)

    if target_date and not nport_filings:
        for file_ref in data.get("filings", {}).get("files", []):
            file_url = f"https://data.sec.gov/submissions/{file_ref['name']}"
            file_resp = sec_get(file_url, headers)
            nport_filings.extend(_extract_nport_filings(file_resp.json(), cik))

    if not nport_filings:
        raise ValueError(f"No NPORT-P filings found for CIK {cik}")

    # Try to find the filing for the correct series by downloading and checking
    matched = _match_filing_to_series(nport_filings, series_id, cik, headers, target_date)
    if matched:
        return matched

    # Last resort: return most recent / closest to target date
    if target_date:
        nport_filings.sort(key=lambda f: abs(_date_diff(f.get("report_date") or f["file_date"], target_date)))
        return nport_filings[0]

    nport_filings.sort(key=lambda f: f["file_date"], reverse=True)
    return nport_filings[0]


def _match_filing_to_series(filings: list, series_id: str, cik: str, headers: dict, target_date: Optional[str]) -> Optional[dict]:
    """Try to match filings to a specific series by checking XML content."""
    candidates = filings[:10]  # Check up to 10 most recent
    if target_date:
        candidates.sort(key=lambda f: abs(_date_diff(f.get("report_date") or f["file_date"], target_date)))
        candidates = candidates[:10]

    for filing in candidates:
        try:
            root = download_nport_xml(filing["cik"], filing["accession"], filing["primary_document"], headers, quiet=True)
            ns = _ns(root)
            gi = root.find(f".//{ns}genInfo")
            if gi is not None:
                found_series = gi.findtext(f"{ns}seriesId", "").strip()
                if found_series == series_id:
                    print(f"  Matched series {series_id} in filing {filing['file_date']}")
                    return filing
        except Exception:
            continue
    return None


def _extract_nport_filings(filings_block: dict, cik: str) -> list:
    """Extract NPORT-P entries from a submissions filings block."""
    results = []
    forms = filings_block.get("form", [])
    filing_dates = filings_block.get("filingDate", [])
    accessions = filings_block.get("accessionNumber", [])
    primary_docs = filings_block.get("primaryDocument", [])
    report_dates = filings_block.get("reportDate", [])
    for i, form in enumerate(forms):
        if form == "NPORT-P":
            results.append({
                "accession": accessions[i],
                "file_date": filing_dates[i],
                "report_date": report_dates[i] if i < len(report_dates) else "",
                "primary_document": primary_docs[i] if i < len(primary_docs) else "",
                "cik": cik,
            })
    return results


def find_all_nport_filings(cik: str, series_id: str, headers: dict) -> list:
    """Return all NPORT-P filings for a specific series via EFTS."""
    print(f"Searching for all NPORT-P filings for series {series_id}...")
    all_filings = []

    try:
        efts_url = (
            f"https://efts.sec.gov/LATEST/search-index"
            f"?q=%22{series_id}%22&forms=NPORT-P"
            f"&dateRange=custom&startdt=2019-01-01&enddt=2030-12-31"
        )
        resp = sec_get(efts_url, headers)
        hits = resp.json().get("hits", {}).get("hits", [])
        for h in hits:
            src = h["_source"]
            cik_from_efts = src["ciks"][0].lstrip("0") if src.get("ciks") else cik
            xsl = src.get("xsl", "")
            primary_doc = f"{xsl}/primary_doc.xml" if xsl else "primary_doc.xml"
            all_filings.append({
                "accession": src["adsh"],
                "file_date": src.get("file_date", ""),
                "report_date": src.get("period_ending", ""),
                "primary_document": primary_doc,
                "cik": cik_from_efts,
            })
    except Exception as e:
        print(f"  EFTS failed ({e}), falling back to submissions API...", file=sys.stderr)
        padded_cik = cik.zfill(10)
        url = f"https://data.sec.gov/submissions/CIK{padded_cik}.json"
        resp = sec_get(url, headers)
        data = resp.json()
        all_filings.extend(_extract_nport_filings(data.get("filings", {}).get("recent", {}), cik))
        for file_ref in data.get("filings", {}).get("files", []):
            file_url = f"https://data.sec.gov/submissions/{file_ref['name']}"
            file_resp = sec_get(file_url, headers)
            all_filings.extend(_extract_nport_filings(file_resp.json(), cik))

    all_filings.sort(key=lambda f: f["file_date"], reverse=True)
    print(f"  Found {len(all_filings)} filings")
    return all_filings


def download_nport_xml(cik: str, accession: str, primary_document: str, headers: dict, quiet: bool = False) -> ET.Element:
    """Download and parse N-PORT XML filing."""
    accession_clean = accession.replace("-", "")

    raw_filename = primary_document.split("/")[-1] if "/" in primary_document else primary_document
    raw_url = f"https://www.sec.gov/Archives/edgar/data/{cik}/{accession_clean}/{raw_filename}"

    if not quiet:
        print(f"Downloading {raw_url}...")
    try:
        resp = sec_get(raw_url, headers)
        root = ET.fromstring(resp.content)
        return root
    except Exception:
        pass

    if "/" in primary_document:
        full_url = f"https://www.sec.gov/Archives/edgar/data/{cik}/{accession_clean}/{primary_document}"
        if not quiet:
            print(f"Fallback: trying {full_url}...")
        try:
            resp = sec_get(full_url, headers)
            root = ET.fromstring(resp.content)
            return root
        except Exception:
            pass

    index_url = f"https://www.sec.gov/Archives/edgar/data/{cik}/{accession_clean}/index.json"
    if not quiet:
        print(f"Fallback: fetching index {index_url}...")
    resp = sec_get(index_url, headers)
    index_data = resp.json()
    for item in index_data.get("directory", {}).get("item", []):
        name = item.get("name", "")
        if name.endswith(".xml") and "primary" in name.lower():
            xml_url = f"https://www.sec.gov/Archives/edgar/data/{cik}/{accession_clean}/{name}"
            resp = sec_get(xml_url, headers)
            return ET.fromstring(resp.content)

    raise ValueError(f"Could not find N-PORT XML for accession {accession}")


def _date_diff(date1: str, date2: str) -> int:
    """Approximate date difference in days for sorting."""
    try:
        d1 = date1.replace("-", "")
        d2 = date2.replace("-", "")
        return abs(int(d1) - int(d2))
    except (ValueError, TypeError):
        return 999999


# ============================================================================
# XML Parsing Helpers
# ============================================================================

def _ns(root: ET.Element) -> str:
    """Extract namespace prefix from root element."""
    tag = root.tag
    if "}" in tag:
        return tag.split("}")[0] + "}"
    return ""


def _ncom_ns(root: ET.Element) -> str:
    """Extract nportcommon namespace."""
    for el in root.iter():
        if "nportcommon" in el.tag:
            return el.tag.split("}")[0] + "}"
    return ""


def _text(el: ET.Element, tag: str, ns: str) -> str:
    """Safely get text of a child element."""
    if el is None:
        return ""
    child = el.find(f"{ns}{tag}")
    if child is not None and child.text:
        return child.text.strip()
    return ""


def _attr(el: ET.Element, tag: str, ns: str, attr_name: str) -> str:
    """Safely get an attribute from a child element."""
    if el is None:
        return ""
    child = el.find(f"{ns}{tag}")
    if child is not None:
        return child.get(attr_name, "")
    return ""


def _deep_text(el: ET.Element, path: str, ns: str) -> str:
    """Safely get text from a deeper path."""
    if el is None:
        return ""
    parts = path.split("/")
    ns_path = "/".join(f"{ns}{p}" if p and not p.startswith(".") else p for p in parts)
    child = el.find(ns_path)
    if child is not None and child.text:
        return child.text.strip()
    return ""


def _float(value: str):
    """Safe float conversion. Returns None on failure."""
    if not value or value in ("N/A", "XXXX", ""):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _to_date(date_str: str):
    """Convert YYYY-MM-DD string to datetime.date for Excel, or return string."""
    if not date_str or date_str in ("N/A", "XXXX", ""):
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return date_str


def _classify_instrument(h: dict) -> str:
    """Return a unified instrument type classification string."""
    ac = h.get("asset_cat", "")
    dt = h.get("deriv_type", "")
    has_debt = h.get("has_debt") == "Y"
    has_repo = h.get("has_repo") == "Y"
    has_deriv = h.get("has_deriv") == "Y"

    if has_repo:
        tc = h.get("repo_trans_cat", "")
        if "reverse" in tc.lower():
            return "Reverse Repo"
        return "Repo"
    if has_deriv:
        if dt == "FWD":
            if ac == "DFE":
                return "FX Forward"
            return "Forward"
        if dt == "FUT":
            if ac == "DIR":
                return "Interest Rate Future"
            if ac == "DE":
                return "Equity Future"
            if ac == "DCO":
                return "Commodity Future"
            return "Future"
        if dt == "SWP":
            if ac == "DCR":
                return "Credit Default Swap"
            if ac == "DIR":
                return "Interest Rate Swap"
            if ac == "DFE":
                return "Cross-Currency Swap"
            if ac == "DE":
                return "Equity Swap"
            if ac == "DCO":
                return "Commodity Swap"
            return "Swap"
        if dt in ("OPT", "SWN", "WAR"):
            label = {"OPT": "Option", "SWN": "Swaption", "WAR": "Warrant"}.get(dt, "Option")
            if ac == "DFE":
                return f"FX {label}"
            if ac == "DIR":
                return f"Interest Rate {label}"
            if ac == "DE":
                return f"Equity {label}"
            return label
        return "Derivative (Other)"
    if has_debt:
        if ac == "ABS-MBS":
            return "MBS"
        if ac == "ABS-CBDO":
            return "CLO/CDO"
        if ac == "ABS-O":
            return "ABS (Other)"
        if ac == "LON":
            return "Loan"
        if ac == "SN":
            return "Structured Note"
        ck = h.get("coupon_kind", "")
        if ck == "Variable":
            return "Floating Rate Note"
        if h.get("is_mandatory_convrtbl") == "Y" or h.get("is_contngt_convrtbl") == "Y":
            return "Convertible Bond"
        return "Fixed Income"
    if ac == "EC":
        return "Equity - Common"
    if ac == "EP":
        return "Equity - Preferred"
    if ac == "STIV":
        return "Money Market / STIV"
    if ac == "LON":
        return "Loan"
    return ASSET_CATEGORY_MAP.get(ac, ac or "Other")


# ============================================================================
# XML Parsing: genInfo
# ============================================================================

def parse_gen_info(root: ET.Element, ns: str) -> dict:
    """Parse genInfo section (registration & fund identity)."""
    gi = root.find(f".//{ns}genInfo")
    if gi is None:
        return {}
    state_el = gi.find(f"{ns}regStateConditional")
    return {
        "reg_name": _text(gi, "regName", ns),
        "reg_file_number": _text(gi, "regFileNumber", ns),
        "reg_cik": _text(gi, "regCik", ns),
        "reg_lei": _text(gi, "regLei", ns),
        "reg_street1": _text(gi, "regStreet1", ns),
        "reg_street2": _text(gi, "regStreet2", ns),
        "reg_city": _text(gi, "regCity", ns),
        "reg_country": state_el.get("regCountry", "") if state_el is not None else "",
        "reg_state": state_el.get("regState", "") if state_el is not None else "",
        "reg_zip": _text(gi, "regZipOrPostalCode", ns),
        "reg_phone": _text(gi, "regPhone", ns),
        "series_name": _text(gi, "seriesName", ns),
        "series_id": _text(gi, "seriesId", ns),
        "series_lei": _text(gi, "seriesLei", ns),
        "rep_pd_end": _text(gi, "repPdEnd", ns),
        "rep_pd_date": _text(gi, "repPdDate", ns),
        "is_final_filing": _text(gi, "isFinalFiling", ns),
    }


# ============================================================================
# XML Parsing: fundInfo
# ============================================================================

def parse_fund_info(root: ET.Element, ns: str) -> dict:
    """Parse fundInfo section (financials, risk, returns, flows, borrowers)."""
    fi = root.find(f".//{ns}fundInfo")
    if fi is None:
        return {}

    result = {}

    # --- Financials ---
    result["financials"] = {
        "tot_assets": _float(_text(fi, "totAssets", ns)),
        "tot_liabs": _float(_text(fi, "totLiabs", ns)),
        "net_assets": _float(_text(fi, "netAssets", ns)),
        "assets_attr_misc_sec": _float(_text(fi, "assetsAttrMiscSec", ns)),
        "assets_invested": _float(_text(fi, "assetsInvested", ns)),
    }

    # --- Borrowing within 1 year ---
    result["borrowing_within_1yr"] = {
        "banks": _float(_text(fi, "amtPayOneYrBanksBorr", ns)),
        "controlled_companies": _float(_text(fi, "amtPayOneYrCtrldComp", ns)),
        "other_affiliates": _float(_text(fi, "amtPayOneYrOthAffil", ns)),
        "other": _float(_text(fi, "amtPayOneYrOther", ns)),
    }

    # --- Borrowing after 1 year ---
    result["borrowing_after_1yr"] = {
        "banks": _float(_text(fi, "amtPayAftOneYrBanksBorr", ns)),
        "controlled_companies": _float(_text(fi, "amtPayAftOneYrCtrldComp", ns)),
        "other_affiliates": _float(_text(fi, "amtPayAftOneYrOthAffil", ns)),
        "other": _float(_text(fi, "amtPayAftOneYrOther", ns)),
    }

    # --- Other financial ---
    result["other_financials"] = {
        "delay_deliv": _float(_text(fi, "delayDeliv", ns)),
        "stand_by_commit": _float(_text(fi, "standByCommit", ns)),
        "liquid_pref": _float(_text(fi, "liquidPref", ns)),
        "cash_not_reported": _float(_text(fi, "cshNotRptdInCorD", ns)),
    }

    # --- Interest rate risk per currency ---
    cur_metrics = []
    for cm in fi.findall(f".//{ns}curMetric"):
        cur_cd = _text(cm, "curCd", ns)
        dv01 = cm.find(f"{ns}intrstRtRiskdv01")
        dv100 = cm.find(f"{ns}intrstRtRiskdv100")
        cur_metrics.append({
            "currency": cur_cd,
            "dv01_3m": _float(dv01.get("period3Mon", "")) if dv01 is not None else None,
            "dv01_1y": _float(dv01.get("period1Yr", "")) if dv01 is not None else None,
            "dv01_5y": _float(dv01.get("period5Yr", "")) if dv01 is not None else None,
            "dv01_10y": _float(dv01.get("period10Yr", "")) if dv01 is not None else None,
            "dv01_30y": _float(dv01.get("period30Yr", "")) if dv01 is not None else None,
            "dv100_3m": _float(dv100.get("period3Mon", "")) if dv100 is not None else None,
            "dv100_1y": _float(dv100.get("period1Yr", "")) if dv100 is not None else None,
            "dv100_5y": _float(dv100.get("period5Yr", "")) if dv100 is not None else None,
            "dv100_10y": _float(dv100.get("period10Yr", "")) if dv100 is not None else None,
            "dv100_30y": _float(dv100.get("period30Yr", "")) if dv100 is not None else None,
        })
    result["currency_metrics"] = cur_metrics

    # --- Credit spread risk ---
    ig = fi.find(f".//{ns}creditSprdRiskInvstGrade")
    nig = fi.find(f".//{ns}creditSprdRiskNonInvstGrade")
    result["credit_spread_risk"] = {
        "inv_grade_3m": _float(ig.get("period3Mon", "")) if ig is not None else None,
        "inv_grade_1y": _float(ig.get("period1Yr", "")) if ig is not None else None,
        "inv_grade_5y": _float(ig.get("period5Yr", "")) if ig is not None else None,
        "inv_grade_10y": _float(ig.get("period10Yr", "")) if ig is not None else None,
        "inv_grade_30y": _float(ig.get("period30Yr", "")) if ig is not None else None,
        "non_inv_grade_3m": _float(nig.get("period3Mon", "")) if nig is not None else None,
        "non_inv_grade_1y": _float(nig.get("period1Yr", "")) if nig is not None else None,
        "non_inv_grade_5y": _float(nig.get("period5Yr", "")) if nig is not None else None,
        "non_inv_grade_10y": _float(nig.get("period10Yr", "")) if nig is not None else None,
        "non_inv_grade_30y": _float(nig.get("period30Yr", "")) if nig is not None else None,
    }

    # --- Borrowers ---
    borrowers = []
    for b in fi.findall(f".//{ns}borrower"):
        borrowers.append({
            "name": b.get("name", "") or _text(b, "name", ns),
            "lei": b.get("lei", "") or _text(b, "lei", ns),
            "aggregate_value": _float(b.get("aggrVal", "") or _text(b, "aggrVal", ns)),
        })
    result["borrowers"] = borrowers

    # --- Monthly total returns ---
    monthly_returns = []
    for mtr in fi.findall(f".//{ns}monthlyTotReturn"):
        monthly_returns.append({
            "class_id": mtr.get("classId", ""),
            "return_month1": _float(mtr.get("rtn1", "")),
            "return_month2": _float(mtr.get("rtn2", "")),
            "return_month3": _float(mtr.get("rtn3", "")),
        })
    result["monthly_total_returns"] = monthly_returns

    # --- Other monthly returns (non-derivative) ---
    result["other_monthly_returns"] = {}
    for tag, key in [("othMon1", "month1"), ("othMon2", "month2"), ("othMon3", "month3")]:
        el = fi.find(f".//{ns}{tag}")
        if el is not None:
            result["other_monthly_returns"][key] = {
                "realized": _float(el.get("netRealizedGain", "")),
                "unrealized": _float(el.get("netUnrealizedAppr", "")),
            }

    # --- Derivative returns by category ---
    deriv_returns = {}
    ret_cats = fi.find(f".//{ns}monthlyReturnCats")
    if ret_cats is not None:
        for elem_name, label in CONTRACT_TYPE_ELEMENTS:
            ct_el = ret_cats.find(f"{ns}{elem_name}")
            if ct_el is None:
                continue
            cat_data = {
                "label": label,
                "mon1_realized": None, "mon1_unrealized": None,
                "mon2_realized": None, "mon2_unrealized": None,
                "mon3_realized": None, "mon3_unrealized": None,
                "instruments": {},
            }
            for mon_tag, mon_key in [("mon1", "mon1"), ("mon2", "mon2"), ("mon3", "mon3")]:
                mon_el = ct_el.find(f"{ns}{mon_tag}")
                if mon_el is not None:
                    cat_data[f"{mon_key}_realized"] = _float(mon_el.get("netRealizedGain", ""))
                    cat_data[f"{mon_key}_unrealized"] = _float(mon_el.get("netUnrealizedAppr", ""))

            for instr_elem, instr_label in INSTRUMENT_CATEGORIES:
                instr_el = ct_el.find(f"{ns}{instr_elem}")
                if instr_el is None:
                    continue
                instr_data = {}
                for mon_tag, mon_key in [("instrMon1", "mon1"), ("instrMon2", "mon2"), ("instrMon3", "mon3")]:
                    m = instr_el.find(f"{ns}{mon_tag}")
                    if m is not None:
                        instr_data[f"{mon_key}_realized"] = _float(m.get("netRealizedGain", ""))
                        instr_data[f"{mon_key}_unrealized"] = _float(m.get("netUnrealizedAppr", ""))
                cat_data["instruments"][instr_label] = instr_data

            deriv_returns[label] = cat_data
    result["derivative_returns"] = deriv_returns

    # --- Flow information ---
    flow_info = {}
    for mon_tag, mon_key in [("mon1Flow", "month1"), ("mon2Flow", "month2"), ("mon3Flow", "month3")]:
        flow_el = fi.find(f".//{ns}{mon_tag}")
        if flow_el is not None:
            flow_info[mon_key] = {
                "sales": _float(flow_el.get("sales", "")),
                "redemption": _float(flow_el.get("redemption", "")),
                "reinvestment": _float(flow_el.get("reinvestment", "")),
            }
    result["flow_info"] = flow_info

    # --- VaR info ---
    var_info = {}
    fdi = fi.find(f".//{ns}fundsDesignatedInfo")
    if fdi is not None:
        var_info["designated_index_name"] = _text(fdi, "nameDesignatedIndex", ns)
        var_info["index_identifier"] = _text(fdi, "indexIdentifier", ns)
    result["var_info"] = var_info

    # --- Non-cash collateral flag ---
    result["is_non_cash_collateral"] = _text(fi, "isNonCashCollateral", ns)

    return result


# ============================================================================
# XML Parsing: Signature
# ============================================================================

def parse_signature(root: ET.Element, ns: str) -> dict:
    """Parse signature section."""
    ncom = _ncom_ns(root)
    sig = root.find(f".//{ns}signature")
    if sig is None:
        return {}
    return {
        "date_signed": sig.findtext(f"{ncom}dateSigned", "").strip() if ncom else "",
        "name_of_applicant": sig.findtext(f"{ncom}nameOfApplicant", "").strip() if ncom else "",
        "signature": sig.findtext(f"{ncom}signature", "").strip() if ncom else "",
        "signer_name": sig.findtext(f"{ncom}signerName", "").strip() if ncom else "",
        "signer_title": sig.findtext(f"{ncom}title", "").strip() if ncom else "",
    }


# ============================================================================
# XML Parsing: Holdings (invstOrSec) — ALL instrument types
# ============================================================================

def parse_holdings(root: ET.Element, ns: str) -> list:
    """Parse all invstOrSec elements into flat dicts."""
    holdings = []
    for inv in root.findall(f".//{ns}invstOrSec"):
        # Skip nested derivative holdings (inside descRefInstrmnt)
        parent_tags = []
        parent = inv
        # Simple depth check: if this invstOrSec is inside nestedDerivInfo, skip
        # We handle nested derivatives inline
        h = _parse_one_holding(inv, ns)
        h["_raw_xml"] = inv  # stash for unmapped-field detection; excluded from sheets
        holdings.append(h)
    return holdings


def _parse_one_holding(inv: ET.Element, ns: str) -> dict:
    """Parse a single invstOrSec element — exhaustive field extraction."""
    h = {}

    # --- Common identifiers ---
    h["name"] = _text(inv, "name", ns)
    h["lei"] = _text(inv, "lei", ns)
    h["title"] = _text(inv, "title", ns)
    h["cusip"] = _text(inv, "cusip", ns)

    # Identifiers block (handle multiple)
    ids = inv.find(f"{ns}identifiers")
    h["isin"] = ""
    h["ticker"] = ""
    h["other_ids"] = ""
    h["other_id_types"] = ""
    if ids is not None:
        isin_el = ids.find(f"{ns}isin")
        if isin_el is not None:
            h["isin"] = isin_el.get("value", "")
        ticker_el = ids.find(f"{ns}ticker")
        if ticker_el is not None:
            h["ticker"] = ticker_el.get("value", "")
        # Collect ALL other identifiers
        other_ids = []
        other_types = []
        for other_el in ids.findall(f"{ns}other"):
            val = other_el.get("value", "")
            desc = other_el.get("otherDesc", "")
            if val:
                other_ids.append(val)
            if desc:
                other_types.append(desc)
        h["other_ids"] = "; ".join(other_ids)
        h["other_id_types"] = "; ".join(other_types)

    # --- Quantity ---
    h["balance"] = _text(inv, "balance", ns)
    h["units"] = _text(inv, "units", ns)
    h["desc_oth_units"] = _text(inv, "descOthUnits", ns)

    # --- Currency (simple or conditional) ---
    h["currency"] = _text(inv, "curCd", ns)
    h["exchange_rate"] = ""
    if not h["currency"] or h["currency"] == "N/A":
        cc = inv.find(f"{ns}currencyConditional")
        if cc is not None:
            h["currency"] = cc.get("curCd", "")
            h["exchange_rate"] = cc.get("exchangeRt", "")

    # --- Valuation ---
    h["value_usd"] = _text(inv, "valUSD", ns)
    h["pct_val"] = _text(inv, "pctVal", ns)
    h["payoff_profile"] = _text(inv, "payoffProfile", ns)

    # --- Asset category (simple or conditional) ---
    h["asset_cat"] = _text(inv, "assetCat", ns)
    h["asset_cat_desc"] = ""
    if not h["asset_cat"]:
        ac = inv.find(f"{ns}assetConditional")
        if ac is not None:
            h["asset_cat"] = ac.get("assetCat", "")
            h["asset_cat_desc"] = ac.get("desc", "")

    # --- Issuer category (simple or conditional) ---
    h["issuer_cat"] = _text(inv, "issuerCat", ns)
    h["issuer_cat_desc"] = ""
    if not h["issuer_cat"]:
        ic = inv.find(f"{ns}issuerConditional")
        if ic is not None:
            h["issuer_cat"] = ic.get("issuerCat", "")
            h["issuer_cat_desc"] = ic.get("desc", "")

    # --- Country / State ---
    h["inv_country"] = _text(inv, "invCountry", ns)
    h["inv_state"] = ""
    inv_state_el = inv.find(f"{ns}invStateConditional")
    if inv_state_el is not None:
        h["inv_country"] = inv_state_el.get("invCountry", h["inv_country"])
        h["inv_state"] = inv_state_el.get("invState", "")

    # --- Other common fields ---
    h["is_restricted_sec"] = _text(inv, "isRestrictedSec", ns)
    h["fair_val_level"] = _text(inv, "fairValLevel", ns)

    # --- Debt security fields ---
    _parse_debt_sec(inv, ns, h)

    # --- Repurchase agreement fields ---
    _parse_repo(inv, ns, h)

    # --- Derivative fields ---
    _parse_derivative(inv, ns, h)

    # --- Security lending fields ---
    _parse_security_lending(inv, ns, h)

    # Unified instrument classification (derived)
    h["instrument_type"] = _classify_instrument(h)

    return h


def _parse_debt_sec(inv: ET.Element, ns: str, h: dict):
    """Parse debtSec child — comprehensive."""
    debt = inv.find(f"{ns}debtSec")
    h["has_debt"] = "Y" if debt is not None else "N"
    h["maturity_dt"] = ""
    h["coupon_kind"] = ""
    h["annualized_rt"] = ""
    h["is_default"] = ""
    h["are_intrst_pmnts_in_arrs"] = ""
    h["is_paid_kind"] = ""
    h["is_mandatory_convrtbl"] = ""
    h["is_contngt_convrtbl"] = ""
    h["debt_ref_name"] = ""
    h["debt_ref_title"] = ""
    h["debt_ref_cusip"] = ""
    h["debt_ref_isin"] = ""
    h["debt_ref_currency"] = ""
    h["conv_ratio"] = ""
    h["conv_currency"] = ""
    h["debt_delta"] = ""
    h["reset_tenors"] = ""  # Semicolon-delimited if multiple

    if debt is None:
        return

    h["maturity_dt"] = _text(debt, "maturityDt", ns)
    h["coupon_kind"] = _text(debt, "couponKind", ns)
    h["annualized_rt"] = _text(debt, "annualizedRt", ns)
    h["is_default"] = _text(debt, "isDefault", ns)
    h["are_intrst_pmnts_in_arrs"] = _text(debt, "areIntrstPmntsInArrs", ns)
    h["is_paid_kind"] = _text(debt, "isPaidKind", ns)
    h["is_mandatory_convrtbl"] = _text(debt, "isMandatoryConvrtbl", ns)
    h["is_contngt_convrtbl"] = _text(debt, "isContngtConvrtbl", ns)
    h["debt_delta"] = _text(debt, "delta", ns)

    # Reference instruments (handle multiple via dbtSecRefInstruments)
    ref_names = []
    ref_titles = []
    ref_cusips = []
    ref_isins = []
    ref_curs = []
    for ri in debt.findall(f".//{ns}dbtSecRefInstrument"):
        n = _text(ri, "name", ns)
        t = _text(ri, "title", ns)
        c = _text(ri, "curCd", ns)
        if n:
            ref_names.append(n)
        if t:
            ref_titles.append(t)
        if c:
            ref_curs.append(c)
        # Identifiers within ref instrument
        ri_ids = ri.find(f"{ns}identifiers")
        if ri_ids is not None:
            isin_el = ri_ids.find(f"{ns}isin")
            if isin_el is not None:
                ref_isins.append(isin_el.get("value", ""))
            cusip_el = ri_ids.find(f"{ns}cusip")
            if cusip_el is not None:
                val = cusip_el.get("value", "") or (cusip_el.text.strip() if cusip_el.text else "")
                if val:
                    ref_cusips.append(val)
    h["debt_ref_name"] = "; ".join(ref_names)
    h["debt_ref_title"] = "; ".join(ref_titles)
    h["debt_ref_cusip"] = "; ".join(ref_cusips)
    h["debt_ref_isin"] = "; ".join(ref_isins)
    h["debt_ref_currency"] = "; ".join(ref_curs)

    # Currency / conversion info (handle multiple via currencyInfos)
    conv_ratios = []
    conv_curs = []
    for ci in debt.findall(f".//{ns}currencyInfo"):
        cr = ci.get("convRatio", "")
        cc = ci.get("curCd", "")
        if cr:
            conv_ratios.append(cr)
        if cc:
            conv_curs.append(cc)
    h["conv_ratio"] = "; ".join(conv_ratios)
    h["conv_currency"] = "; ".join(conv_curs)

    # Rate reset tenors (handle ALL via rtResetTenors)
    tenors = []
    for rt in debt.findall(f".//{ns}rtResetTenor"):
        tenor_str = f"{rt.get('rateTenor', '')} {rt.get('rateTenorUnit', '')}".strip()
        reset_str = f"reset {rt.get('resetDt', '')} {rt.get('resetDtUnit', '')}".strip()
        tenors.append(f"{tenor_str} ({reset_str})")
    h["reset_tenors"] = "; ".join(tenors)


def _parse_repo(inv: ET.Element, ns: str, h: dict):
    """Parse repurchaseAgrmt child — covers both repo and reverse repo."""
    repo = inv.find(f"{ns}repurchaseAgrmt")
    h["has_repo"] = "Y" if repo is not None else "N"
    h["repo_trans_cat"] = ""
    h["repo_is_cleared"] = ""
    h["repo_counterparty_name"] = ""
    h["repo_counterparty_lei"] = ""
    h["repo_is_tri_party"] = ""
    h["repo_rate"] = ""
    h["repo_maturity_dt"] = ""
    h["repo_collateral_principal"] = ""
    h["repo_collateral_principal_cur"] = ""
    h["repo_collateral_value"] = ""
    h["repo_collateral_value_cur"] = ""
    h["repo_collateral_category"] = ""
    h["repo_collateral_category_desc"] = ""

    if repo is None:
        return

    h["repo_trans_cat"] = _text(repo, "transCat", ns)
    h["repo_is_tri_party"] = _text(repo, "isTriParty", ns)
    h["repo_rate"] = _text(repo, "repurchaseRt", ns)
    h["repo_maturity_dt"] = _text(repo, "maturityDt", ns)

    # Clearing info
    cleared_el = repo.find(f"{ns}notClearedCentCparty")
    if cleared_el is not None:
        h["repo_is_cleared"] = cleared_el.get("isCleared", "")
        # Counterparties within notClearedCentCparty
        cp_names = []
        cp_leis = []
        for cp in cleared_el.findall(f".//{ns}counterpartyInfo"):
            n = cp.get("name", "")
            l = cp.get("lei", "")
            if n:
                cp_names.append(n)
            if l:
                cp_leis.append(l)
        h["repo_counterparty_name"] = "; ".join(cp_names)
        h["repo_counterparty_lei"] = "; ".join(cp_leis)

    # Collateral (handle multiple collateral items)
    principals = []
    principal_curs = []
    coll_vals = []
    coll_curs = []
    coll_cats = []
    coll_descs = []
    for rc in repo.findall(f".//{ns}repurchaseCollateral"):
        pa = _text(rc, "principalAmt", ns)
        pc = _text(rc, "principalCd", ns)
        cv = _text(rc, "collateralVal", ns)
        cc = _text(rc, "collateralCd", ns)
        ic = _text(rc, "invstCat", ns)
        ic_desc = ""
        # Also check invstCatConditional
        icc = rc.find(f"{ns}invstCatConditional")
        if icc is not None:
            if not ic:
                ic = icc.get("invstCat", "")
            ic_desc = icc.get("desc", "")
        if pa:
            principals.append(pa)
        if pc:
            principal_curs.append(pc)
        if cv:
            coll_vals.append(cv)
        if cc:
            coll_curs.append(cc)
        if ic:
            coll_cats.append(ic)
        if ic_desc:
            coll_descs.append(ic_desc)
    h["repo_collateral_principal"] = "; ".join(principals)
    h["repo_collateral_principal_cur"] = "; ".join(principal_curs)
    h["repo_collateral_value"] = "; ".join(coll_vals)
    h["repo_collateral_value_cur"] = "; ".join(coll_curs)
    h["repo_collateral_category"] = "; ".join(coll_cats)
    h["repo_collateral_category_desc"] = "; ".join(coll_descs)


def _parse_derivative(inv: ET.Element, ns: str, h: dict):
    """Parse derivativeInfo child — all derivative types."""
    deriv = inv.find(f"{ns}derivativeInfo")
    h["has_deriv"] = "Y" if deriv is not None else "N"
    h["deriv_type"] = ""
    h["deriv_type_desc"] = ""
    h["counterparty_name"] = ""
    h["counterparty_lei"] = ""
    h["unrealized_appr"] = ""

    # Forward fields
    h["fwd_amt_cur_sold"] = ""
    h["fwd_cur_sold"] = ""
    h["fwd_amt_cur_pur"] = ""
    h["fwd_cur_pur"] = ""
    h["fwd_settlement_dt"] = ""

    # Future fields
    h["fut_payoff_prof"] = ""
    h["fut_exp_date"] = ""
    h["fut_notional_amt"] = ""
    h["fut_notional_cur"] = ""

    # Swap fields
    h["swap_flag"] = ""
    h["swap_rec_fixed_or_floating"] = ""
    h["swap_fixed_rec_rt"] = ""
    h["swap_fixed_rec_amt"] = ""
    h["swap_fixed_rec_cur"] = ""
    h["swap_float_rec_index"] = ""
    h["swap_float_rec_spread"] = ""
    h["swap_float_rec_pmnt_amt"] = ""
    h["swap_float_rec_cur"] = ""
    h["swap_float_rec_reset_tenor"] = ""
    h["swap_pay_fixed_or_floating"] = ""
    h["swap_fixed_pay_rt"] = ""
    h["swap_fixed_pay_amt"] = ""
    h["swap_fixed_pay_cur"] = ""
    h["swap_float_pay_index"] = ""
    h["swap_float_pay_spread"] = ""
    h["swap_float_pay_pmnt_amt"] = ""
    h["swap_float_pay_cur"] = ""
    h["swap_float_pay_reset_tenor"] = ""
    h["swap_other_rec_desc"] = ""
    h["swap_other_pay_desc"] = ""
    h["swap_termination_dt"] = ""
    h["swap_upfront_pmnt"] = ""
    h["swap_pmnt_cur"] = ""
    h["swap_upfront_rcpt"] = ""
    h["swap_rcpt_cur"] = ""
    h["swap_notional_amt"] = ""
    h["swap_notional_cur"] = ""

    # Option/Swaption/Warrant fields
    h["opt_put_or_call"] = ""
    h["opt_written_or_pur"] = ""
    h["opt_share_no"] = ""
    h["opt_exercise_price"] = ""
    h["opt_exercise_price_cur"] = ""
    h["opt_exp_dt"] = ""
    h["opt_delta"] = ""
    h["opt_nested_deriv_type"] = ""
    h["opt_nested_deriv_title"] = ""
    # Nested derivative full details (derivAddlInfo inside nestedDerivInfo)
    h["opt_nested_deriv_name"] = ""
    h["opt_nested_deriv_lei"] = ""
    h["opt_nested_deriv_cusip"] = ""
    h["opt_nested_deriv_isin"] = ""
    h["opt_nested_deriv_other_id"] = ""
    h["opt_nested_deriv_other_desc"] = ""
    h["opt_nested_deriv_balance"] = ""
    h["opt_nested_deriv_units"] = ""
    h["opt_nested_deriv_currency"] = ""
    h["opt_nested_deriv_exchange_rt"] = ""
    h["opt_nested_deriv_val_usd"] = ""
    h["opt_nested_deriv_pct_val"] = ""
    h["opt_nested_deriv_asset_cat"] = ""
    h["opt_nested_deriv_issuer_cat"] = ""
    h["opt_nested_deriv_issuer_desc"] = ""
    h["opt_nested_deriv_inv_country"] = ""
    # Nested derivative-specific fields (from the derivative element itself)
    h["opt_nested_deriv_cat"] = ""
    h["opt_nested_deriv_counterparty"] = ""
    h["opt_nested_deriv_counterparty_lei"] = ""
    h["opt_nested_deriv_settlement_dt"] = ""
    h["opt_nested_deriv_fwd_amt_sold"] = ""
    h["opt_nested_deriv_fwd_cur_sold"] = ""
    h["opt_nested_deriv_fwd_amt_pur"] = ""
    h["opt_nested_deriv_fwd_cur_pur"] = ""
    h["opt_nested_deriv_swap_rec_index"] = ""
    h["opt_nested_deriv_swap_rec_rt"] = ""
    h["opt_nested_deriv_swap_pay_index"] = ""
    h["opt_nested_deriv_swap_pay_rt"] = ""
    h["opt_nested_deriv_swap_termination_dt"] = ""
    h["opt_nested_deriv_swap_notional"] = ""
    h["opt_nested_deriv_swap_notional_cur"] = ""
    h["opt_nested_deriv_swap_rec_fixed_or_float"] = ""
    h["opt_nested_deriv_swap_pay_fixed_or_float"] = ""
    h["opt_nested_deriv_swap_rec_reset_tenor"] = ""
    h["opt_nested_deriv_swap_pay_reset_tenor"] = ""

    # Reference instrument (shared across derivative types)
    h["deriv_ref_index_name"] = ""
    h["deriv_ref_index_id"] = ""
    h["deriv_ref_narrative"] = ""
    h["deriv_ref_issuer_name"] = ""
    h["deriv_ref_issue_title"] = ""
    h["deriv_ref_cusip"] = ""
    h["deriv_ref_isin"] = ""
    h["deriv_ref_other_id"] = ""
    h["deriv_ref_other_desc"] = ""

    if deriv is None:
        return

    # Determine derivative type
    for child in deriv:
        tag = child.tag.replace(ns, "") if ns else child.tag

        if tag == "fwdDeriv":
            h["deriv_type"] = child.get("derivCat", "FWD")
            _parse_counterparties(child, ns, h)
            h["fwd_amt_cur_sold"] = _text(child, "amtCurSold", ns)
            h["fwd_cur_sold"] = _text(child, "curSold", ns)
            h["fwd_amt_cur_pur"] = _text(child, "amtCurPur", ns)
            h["fwd_cur_pur"] = _text(child, "curPur", ns)
            h["fwd_settlement_dt"] = _text(child, "settlementDt", ns)
            h["unrealized_appr"] = _text(child, "unrealizedAppr", ns)
            _parse_ref_instrument(child, ns, h)

        elif tag == "futrDeriv":
            h["deriv_type"] = child.get("derivCat", "FUT")
            _parse_counterparties(child, ns, h)
            h["fut_payoff_prof"] = _text(child, "payOffProf", ns)
            h["fut_exp_date"] = _text(child, "expDate", ns)
            h["fut_notional_amt"] = _text(child, "notionalAmt", ns)
            h["fut_notional_cur"] = _text(child, "curCd", ns)
            h["unrealized_appr"] = _text(child, "unrealizedAppr", ns)
            _parse_ref_instrument(child, ns, h)

        elif tag == "swapDeriv":
            h["deriv_type"] = child.get("derivCat", "SWP")
            _parse_counterparties(child, ns, h)
            h["swap_flag"] = _text(child, "swapFlag", ns)

            # Fixed receive leg
            fr = child.find(f"{ns}fixedRecDesc")
            if fr is not None:
                h["swap_rec_fixed_or_floating"] = fr.get("fixedOrFloating", "Fixed")
                h["swap_fixed_rec_rt"] = fr.get("fixedRt", "")
                h["swap_fixed_rec_amt"] = fr.get("amount", "")
                h["swap_fixed_rec_cur"] = fr.get("curCd", "")
            # Floating receive leg
            flr = child.find(f"{ns}floatingRecDesc")
            if flr is not None:
                h["swap_rec_fixed_or_floating"] = flr.get("fixedOrFloating", "Floating")
                h["swap_float_rec_index"] = flr.get("floatingRtIndex", "")
                h["swap_float_rec_spread"] = flr.get("floatingRtSpread", "")
                h["swap_float_rec_pmnt_amt"] = flr.get("pmntAmt", "")
                h["swap_float_rec_cur"] = flr.get("curCd", "")
                # Reset tenors within floating receive
                tenors = []
                for rt in flr.findall(f".//{ns}rtResetTenor"):
                    t = f"{rt.get('rateTenor', '')} {rt.get('rateTenorUnit', '')}".strip()
                    r = f"reset {rt.get('resetDt', '')} {rt.get('resetDtUnit', '')}".strip()
                    tenors.append(f"{t} ({r})")
                h["swap_float_rec_reset_tenor"] = "; ".join(tenors)
            # Other receive leg (CDS single-leg etc.)
            orec = child.find(f"{ns}otherRecDesc")
            if orec is not None:
                h["swap_rec_fixed_or_floating"] = orec.get("fixedOrFloating", "Other")
                h["swap_other_rec_desc"] = (orec.text or "").strip()

            # Fixed pay leg
            fp = child.find(f"{ns}fixedPmntDesc")
            if fp is not None:
                h["swap_pay_fixed_or_floating"] = fp.get("fixedOrFloating", "Fixed")
                h["swap_fixed_pay_rt"] = fp.get("fixedRt", "")
                h["swap_fixed_pay_amt"] = fp.get("amount", "")
                h["swap_fixed_pay_cur"] = fp.get("curCd", "")
            # Floating pay leg
            flp = child.find(f"{ns}floatingPmntDesc")
            if flp is not None:
                h["swap_pay_fixed_or_floating"] = flp.get("fixedOrFloating", "Floating")
                h["swap_float_pay_index"] = flp.get("floatingRtIndex", "")
                h["swap_float_pay_spread"] = flp.get("floatingRtSpread", "")
                h["swap_float_pay_pmnt_amt"] = flp.get("pmntAmt", "")
                h["swap_float_pay_cur"] = flp.get("curCd", "")
                tenors = []
                for rt in flp.findall(f".//{ns}rtResetTenor"):
                    t = f"{rt.get('rateTenor', '')} {rt.get('rateTenorUnit', '')}".strip()
                    r = f"reset {rt.get('resetDt', '')} {rt.get('resetDtUnit', '')}".strip()
                    tenors.append(f"{t} ({r})")
                h["swap_float_pay_reset_tenor"] = "; ".join(tenors)
            # Other pay leg
            opay = child.find(f"{ns}otherPmntDesc")
            if opay is not None:
                h["swap_pay_fixed_or_floating"] = opay.get("fixedOrFloating", "Other")
                h["swap_other_pay_desc"] = (opay.text or "").strip()

            h["swap_termination_dt"] = _text(child, "terminationDt", ns)
            h["swap_upfront_pmnt"] = _text(child, "upfrontPmnt", ns)
            h["swap_pmnt_cur"] = _text(child, "pmntCurCd", ns)
            h["swap_upfront_rcpt"] = _text(child, "upfrontRcpt", ns)
            h["swap_rcpt_cur"] = _text(child, "rcptCurCd", ns)
            h["swap_notional_amt"] = _text(child, "notionalAmt", ns)
            h["swap_notional_cur"] = _text(child, "curCd", ns)
            h["unrealized_appr"] = _text(child, "unrealizedAppr", ns)

            # Swap also has forward-like fields for CCS (cross-currency swaps)
            amt_sold = _text(child, "amtCurSold", ns)
            if amt_sold:
                h["fwd_amt_cur_sold"] = amt_sold
                h["fwd_cur_sold"] = _text(child, "curSold", ns)
                h["fwd_amt_cur_pur"] = _text(child, "amtCurPur", ns)
                h["fwd_cur_pur"] = _text(child, "curPur", ns)
                h["fwd_settlement_dt"] = _text(child, "settlementDt", ns)

            _parse_ref_instrument(child, ns, h)

        elif tag == "optionSwaptionWarrantDeriv":
            h["deriv_type"] = child.get("derivCat", "OPT")
            _parse_counterparties(child, ns, h)
            h["opt_put_or_call"] = _text(child, "putOrCall", ns)
            h["opt_written_or_pur"] = _text(child, "writtenOrPur", ns)
            h["opt_share_no"] = _text(child, "shareNo", ns)
            h["opt_exercise_price"] = _text(child, "exercisePrice", ns)
            h["opt_exercise_price_cur"] = _text(child, "exercisePriceCurCd", ns)
            h["opt_exp_dt"] = _text(child, "expDt", ns)
            h["opt_delta"] = _text(child, "delta", ns)
            h["unrealized_appr"] = _text(child, "unrealizedAppr", ns)

            # Nested derivative info (e.g. swaption on a forward)
            nested = child.find(f".//{ns}nestedDerivInfo")
            if nested is not None:
                for nc in nested:
                    ntag = nc.tag.replace(ns, "") if ns else nc.tag
                    h["opt_nested_deriv_type"] = ntag
                    h["opt_nested_deriv_cat"] = nc.get("derivCat", "")
                    # derivAddlInfo — full holding-like info for the referenced derivative
                    addl = nc.find(f"{ns}derivAddlInfo")
                    if addl is not None:
                        h["opt_nested_deriv_name"] = _text(addl, "name", ns)
                        h["opt_nested_deriv_lei"] = _text(addl, "lei", ns)
                        h["opt_nested_deriv_title"] = _text(addl, "title", ns)
                        h["opt_nested_deriv_cusip"] = _text(addl, "cusip", ns)
                        # Identifiers (isin, other)
                        ids_el = addl.find(f"{ns}identifiers")
                        if ids_el is not None:
                            isin_el = ids_el.find(f"{ns}isin")
                            if isin_el is not None:
                                h["opt_nested_deriv_isin"] = isin_el.get("value", "")
                            other_el = ids_el.find(f"{ns}other")
                            if other_el is not None:
                                h["opt_nested_deriv_other_id"] = other_el.get("value", "")
                                h["opt_nested_deriv_other_desc"] = other_el.get("otherDesc", "")
                        h["opt_nested_deriv_balance"] = _text(addl, "balance", ns)
                        h["opt_nested_deriv_units"] = _text(addl, "units", ns)
                        cur_cond = addl.find(f"{ns}currencyConditional")
                        if cur_cond is not None:
                            h["opt_nested_deriv_currency"] = cur_cond.get("curCd", "")
                            h["opt_nested_deriv_exchange_rt"] = cur_cond.get("exchangeRt", "")
                        h["opt_nested_deriv_val_usd"] = _text(addl, "valUSD", ns)
                        h["opt_nested_deriv_pct_val"] = _text(addl, "pctVal", ns)
                        h["opt_nested_deriv_asset_cat"] = _text(addl, "assetCat", ns)
                        issuer_cond = addl.find(f"{ns}issuerConditional")
                        if issuer_cond is not None:
                            h["opt_nested_deriv_issuer_cat"] = issuer_cond.get("issuerCat", "")
                            h["opt_nested_deriv_issuer_desc"] = issuer_cond.get("desc", "")
                        h["opt_nested_deriv_inv_country"] = _text(addl, "invCountry", ns)
                    # Nested derivative counterparties
                    n_cps = nc.findall(f"{ns}counterparties")
                    n_names, n_leis = [], []
                    for ncp in n_cps:
                        nn = _text(ncp, "counterpartyName", ns)
                        nl = _text(ncp, "counterpartyLei", ns)
                        if nn:
                            n_names.append(nn)
                        if nl:
                            n_leis.append(nl)
                    h["opt_nested_deriv_counterparty"] = "; ".join(n_names)
                    h["opt_nested_deriv_counterparty_lei"] = "; ".join(n_leis)
                    # Nested forward fields
                    if ntag == "fwdDeriv":
                        h["opt_nested_deriv_settlement_dt"] = _text(nc, "settlementDt", ns)
                        h["opt_nested_deriv_fwd_amt_sold"] = _text(nc, "amtCurSold", ns)
                        h["opt_nested_deriv_fwd_cur_sold"] = _text(nc, "curSold", ns)
                        h["opt_nested_deriv_fwd_amt_pur"] = _text(nc, "amtCurPur", ns)
                        h["opt_nested_deriv_fwd_cur_pur"] = _text(nc, "curPur", ns)
                    # Nested swap fields
                    elif ntag == "swapDeriv":
                        h["opt_nested_deriv_swap_termination_dt"] = _text(nc, "terminationDt", ns)
                        h["opt_nested_deriv_swap_notional"] = _text(nc, "notionalAmt", ns)
                        h["opt_nested_deriv_swap_notional_cur"] = _text(nc, "curCd", ns)
                        # Receive leg (floating or fixed)
                        nfr = nc.find(f"{ns}floatingRecDesc")
                        if nfr is not None:
                            h["opt_nested_deriv_swap_rec_fixed_or_float"] = nfr.get("fixedOrFloating", "Floating")
                            h["opt_nested_deriv_swap_rec_index"] = nfr.get("floatingRtIndex", "")
                            h["opt_nested_deriv_swap_rec_rt"] = nfr.get("floatingRtSpread", "")
                            # Reset tenors
                            tenors = []
                            for rt in nfr.findall(f".//{ns}rtResetTenor"):
                                t = f"{rt.get('rateTenor', '')} {rt.get('rateTenorUnit', '')}".strip()
                                r = f"reset {rt.get('resetDt', '')} {rt.get('resetDtUnit', '')}".strip()
                                tenors.append(f"{t} ({r})" if r != "reset" else t)
                            h["opt_nested_deriv_swap_rec_reset_tenor"] = "; ".join(tenors)
                        nfxr = nc.find(f"{ns}fixedRecDesc")
                        if nfxr is not None:
                            h["opt_nested_deriv_swap_rec_fixed_or_float"] = nfxr.get("fixedOrFloating", "Fixed")
                            h["opt_nested_deriv_swap_rec_rt"] = nfxr.get("fixedRt", "")
                        # Pay leg (fixed or floating)
                        nfp = nc.find(f"{ns}fixedPmntDesc")
                        if nfp is not None:
                            h["opt_nested_deriv_swap_pay_fixed_or_float"] = nfp.get("fixedOrFloating", "Fixed")
                            h["opt_nested_deriv_swap_pay_rt"] = nfp.get("fixedRt", "")
                        nflp = nc.find(f"{ns}floatingPmntDesc")
                        if nflp is not None:
                            h["opt_nested_deriv_swap_pay_fixed_or_float"] = nflp.get("fixedOrFloating", "Floating")
                            h["opt_nested_deriv_swap_pay_index"] = nflp.get("floatingRtIndex", "")
                            h["opt_nested_deriv_swap_pay_rt"] = nflp.get("floatingRtSpread", "")
                            # Reset tenors
                            tenors = []
                            for rt in nflp.findall(f".//{ns}rtResetTenor"):
                                t = f"{rt.get('rateTenor', '')} {rt.get('rateTenorUnit', '')}".strip()
                                r = f"reset {rt.get('resetDt', '')} {rt.get('resetDtUnit', '')}".strip()
                                tenors.append(f"{t} ({r})" if r != "reset" else t)
                            h["opt_nested_deriv_swap_pay_reset_tenor"] = "; ".join(tenors)

            _parse_ref_instrument(child, ns, h)

    h["deriv_type_desc"] = DERIV_CATEGORY_MAP.get(h["deriv_type"], h["deriv_type"])


def _parse_counterparties(deriv_el: ET.Element, ns: str, h: dict):
    """Extract counterparty names and LEIs from a derivative element."""
    names = []
    leis = []
    for cp in deriv_el.findall(f"{ns}counterparties"):
        n = _text(cp, "counterpartyName", ns)
        l = _text(cp, "counterpartyLei", ns)
        if n:
            names.append(n)
        if l:
            leis.append(l)
    h["counterparty_name"] = "; ".join(names)
    h["counterparty_lei"] = "; ".join(leis)


def _parse_ref_instrument(deriv_el: ET.Element, ns: str, h: dict):
    """Extract reference instrument info from a derivative element."""
    desc = deriv_el.find(f"{ns}descRefInstrmnt")
    if desc is None:
        desc = deriv_el  # Some filings put ref info directly on deriv element

    # Index basket
    idx = desc.find(f"{ns}indexBasketInfo")
    if idx is not None:
        h["deriv_ref_index_name"] = _text(idx, "indexName", ns)
        h["deriv_ref_index_id"] = _text(idx, "indexIdentifier", ns)
        h["deriv_ref_narrative"] = _text(idx, "narrativeDesc", ns)

    # Other reference instrument
    other = desc.find(f"{ns}otherRefInst")
    if other is not None:
        h["deriv_ref_issuer_name"] = _text(other, "issuerName", ns)
        h["deriv_ref_issue_title"] = _text(other, "issueTitle", ns)
        ref_ids = other.find(f"{ns}identifiers")
        if ref_ids is not None:
            isin_el = ref_ids.find(f"{ns}isin")
            if isin_el is not None:
                h["deriv_ref_isin"] = isin_el.get("value", "")
            cusip_el = ref_ids.find(f"{ns}cusip")
            if cusip_el is not None:
                h["deriv_ref_cusip"] = cusip_el.get("value", "") or (cusip_el.text.strip() if cusip_el.text else "")
            ref_other_ids = []
            ref_other_descs = []
            for oth in ref_ids.findall(f"{ns}other"):
                val = oth.get("value", "")
                desc = oth.get("otherDesc", "")
                if val:
                    ref_other_ids.append(val)
                    if not h["deriv_ref_cusip"]:
                        h["deriv_ref_cusip"] = val
                if desc:
                    ref_other_descs.append(desc)
            h["deriv_ref_other_id"] = "; ".join(ref_other_ids)
            h["deriv_ref_other_desc"] = "; ".join(ref_other_descs)


def _parse_security_lending(inv: ET.Element, ns: str, h: dict):
    """Parse securityLending child — comprehensive."""
    sl = inv.find(f"{ns}securityLending")
    h["sl_is_cash_collateral"] = ""
    h["sl_cash_collateral_val"] = ""
    h["sl_is_non_cash_collateral"] = ""
    h["sl_is_loan_by_fund"] = ""
    h["sl_loan_val"] = ""

    if sl is None:
        return

    # Cash collateral: either simple text or conditional with attributes
    h["sl_is_cash_collateral"] = _text(sl, "isCashCollateral", ns)
    cc = sl.find(f"{ns}cashCollateralCondition")
    if cc is not None:
        h["sl_is_cash_collateral"] = cc.get("isCashCollateral", "")
        h["sl_cash_collateral_val"] = cc.get("cashCollateralVal", "")

    h["sl_is_non_cash_collateral"] = _text(sl, "isNonCashCollateral", ns)

    # Loan by fund: either simple text or conditional
    h["sl_is_loan_by_fund"] = _text(sl, "isLoanByFund", ns)
    lbf = sl.find(f"{ns}loanByFundCondition")
    if lbf is not None:
        h["sl_is_loan_by_fund"] = lbf.get("isLoanByFund", "")
        h["sl_loan_val"] = lbf.get("loanVal", "")


# ============================================================================
# XML Parsing: Explanatory Notes
# ============================================================================

def parse_explanatory_notes(root: ET.Element, ns: str) -> list:
    """Parse all explanatory notes."""
    notes = []
    for note in root.findall(f".//{ns}explntrNote"):
        notes.append({
            "item": note.get("noteItem", ""),
            "note": note.get("note", ""),
        })
    return notes


# ============================================================================
# Excel Writing: Styles & Helpers
# ============================================================================

def create_workbook():
    """Create workbook with shared styles."""
    wb = openpyxl.Workbook()
    styles = {
        "header_fill": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
        "header_font": Font(bold=True, color="FFFFFF", size=10),
        "header_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "title_font": Font(bold=True, size=12),
        "section_font": Font(bold=True, size=11, color="2F5496"),
        "bold_font": Font(bold=True, size=10),
        "normal_font": Font(size=10),
        "money_fmt": "$#,##0.00",
        "pct_fmt": "0.0000",
        "rate_fmt": "0.0000",
        "num_fmt": "#,##0.00",
        "int_fmt": "#,##0",
    }
    return wb, styles


def _write_header_row(ws, headers, styles):
    """Write formatted header row."""
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["header_align"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _auto_width(ws, max_width=45):
    """Auto-size columns based on content."""
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        width = min(max_len + 2, max_width)
        ws.column_dimensions[col_cells[0].column_letter].width = max(width, 8)


def _safe_float_val(val_str):
    """Convert string to float for Excel, return None if empty."""
    if not val_str or val_str in ("N/A", "XXXX", ""):
        return None
    try:
        return float(val_str)
    except (ValueError, TypeError):
        return val_str


# ============================================================================
# Excel Writing: Individual Sheets
# ============================================================================

def write_holdings_sheet(wb, holdings, styles, ticker):
    """Sheet 1: All holdings with common fields."""
    ws = wb.active
    ws.title = f"{ticker} Holdings"

    headers = [
        "Instrument Type",
        "Name", "LEI", "Title", "CUSIP", "ISIN", "Ticker",
        "Other IDs", "Other ID Types",
        "Balance", "Units", "Units Desc",
        "Currency", "Exchange Rate",
        "Value (USD)", "% of NAV",
        "Payoff Profile",
        "Asset Category", "Asset Cat Desc",
        "Issuer Category", "Issuer Cat Desc",
        "Country", "State", "Is Restricted", "Fair Value Level",
        "Has Debt", "Has Derivative", "Deriv Type", "Has Repo",
        "Maturity Date",
        "Is Cash Collateral", "Is Loan By Fund",
    ]
    _write_header_row(ws, headers, styles)

    red_font = Font(color="CC0000", size=10)
    for row_idx, h in enumerate(holdings, 2):
        vals = [
            h["instrument_type"],
            h["name"], h["lei"], h["title"], h["cusip"], h["isin"], h["ticker"],
            h["other_ids"], h["other_id_types"],
            _safe_float_val(h["balance"]), h["units"], UNITS_MAP.get(h["units"], h.get("desc_oth_units", "")),
            h["currency"], _safe_float_val(h["exchange_rate"]),
            _safe_float_val(h["value_usd"]), _safe_float_val(h["pct_val"]),
            h["payoff_profile"],
            h["asset_cat"], ASSET_CATEGORY_MAP.get(h["asset_cat"], h.get("asset_cat_desc", "")),
            h["issuer_cat"], ISSUER_CATEGORY_MAP.get(h["issuer_cat"], h.get("issuer_cat_desc", "")),
            h["inv_country"], h["inv_state"], h["is_restricted_sec"], h["fair_val_level"],
            h["has_debt"], h["has_deriv"], h.get("deriv_type_desc", ""), h["has_repo"],
            _to_date(h.get("maturity_dt", "")),
            h["sl_is_cash_collateral"], h["sl_is_loan_by_fund"],
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
            if col_idx == 15 and isinstance(val, (int, float)):
                cell.number_format = styles["money_fmt"]
                if val < 0:
                    cell.font = red_font
            elif col_idx == 16 and isinstance(val, (int, float)):
                cell.number_format = styles["pct_fmt"]
            elif col_idx == 10 and isinstance(val, (int, float)):
                cell.number_format = styles["num_fmt"]
            elif col_idx == 30 and isinstance(val, date):
                cell.number_format = "YYYY-MM-DD"

    _auto_width(ws)


def write_derivatives_sheet(wb, holdings, styles):
    """Sheet 2: Derivative positions with full details."""
    derivs = [h for h in holdings if h["has_deriv"] == "Y"]
    if not derivs:
        return

    ws = wb.create_sheet("Derivatives")
    headers = [
        "Name", "Title", "CUSIP", "ISIN",
        "Deriv Type", "Deriv Type Desc",
        "Counterparty", "Counterparty LEI",
        "Value (USD)", "% of NAV",
        "Unrealized Appr",
        # Forward / CCS
        "Fwd Amt Sold", "Fwd Cur Sold", "Fwd Amt Purchased", "Fwd Cur Purchased", "Fwd Settlement Date",
        # Future
        "Fut Payoff Profile", "Fut Exp Date", "Fut Notional Amt", "Fut Notional Cur",
        # Swap
        "Swap Flag", "Swap Notional Amt", "Swap Notional Cur",
        "Rec Fixed/Float", "Swap Fixed Rec Rate", "Swap Float Rec Index", "Swap Float Rec Spread", "Swap Float Rec Reset Tenor",
        "Pay Fixed/Float", "Swap Fixed Pay Rate", "Swap Float Pay Index", "Swap Float Pay Spread", "Swap Float Pay Reset Tenor",
        "Swap Other Rec Desc", "Swap Other Pay Desc",
        "Swap Termination Date", "Swap Upfront Pmnt", "Swap Upfront Rcpt",
        # Option/Swaption/Warrant
        "Put/Call", "Written/Purchased",
        "Shares/Contracts", "Exercise Price", "Exercise Price Cur",
        "Opt Exp Date", "Delta",
        "Nested Deriv Type", "Nested Deriv Title",
        "Nested Deriv Name", "Nested Deriv LEI", "Nested Deriv CUSIP", "Nested Deriv ISIN",
        "Nested Deriv Other ID", "Nested Deriv Other Desc",
        "Nested Deriv Balance", "Nested Deriv Units", "Nested Deriv Currency", "Nested Deriv Exchange Rate",
        "Nested Deriv Value (USD)", "Nested Deriv % NAV",
        "Nested Deriv Asset Cat", "Nested Deriv Issuer Cat", "Nested Deriv Issuer Desc",
        "Nested Deriv Country",
        "Nested Deriv Cat", "Nested Deriv Counterparty", "Nested Deriv Counterparty LEI",
        "Nested Deriv Settlement Date",
        "Nested Fwd Amt Sold", "Nested Fwd Cur Sold", "Nested Fwd Amt Purchased", "Nested Fwd Cur Purchased",
        "Nested Swap Rec Fixed/Float", "Nested Swap Rec Index", "Nested Swap Rec Rate", "Nested Swap Rec Reset Tenor",
        "Nested Swap Pay Fixed/Float", "Nested Swap Pay Index", "Nested Swap Pay Rate", "Nested Swap Pay Reset Tenor",
        "Nested Swap Termination Date", "Nested Swap Notional", "Nested Swap Notional Cur",
        # Reference instrument
        "Ref Index Name", "Ref Index ID", "Ref Narrative",
        "Ref Issuer Name", "Ref Issue Title", "Ref CUSIP", "Ref ISIN",
        "Ref Other ID", "Ref Other Desc",
    ]
    _write_header_row(ws, headers, styles)

    for row_idx, h in enumerate(derivs, 2):
        vals = [
            h["name"], h["title"], h["cusip"], h["isin"],
            h["deriv_type"], h["deriv_type_desc"],
            h["counterparty_name"], h["counterparty_lei"],
            _safe_float_val(h["value_usd"]), _safe_float_val(h["pct_val"]),
            _safe_float_val(h["unrealized_appr"]),
            _safe_float_val(h["fwd_amt_cur_sold"]), h["fwd_cur_sold"],
            _safe_float_val(h["fwd_amt_cur_pur"]), h["fwd_cur_pur"],
            h["fwd_settlement_dt"],
            h["fut_payoff_prof"], h["fut_exp_date"],
            _safe_float_val(h["fut_notional_amt"]), h["fut_notional_cur"],
            h["swap_flag"],
            _safe_float_val(h["swap_notional_amt"]), h["swap_notional_cur"],
            h["swap_rec_fixed_or_floating"],
            _safe_float_val(h["swap_fixed_rec_rt"]),
            h["swap_float_rec_index"], _safe_float_val(h["swap_float_rec_spread"]),
            h["swap_float_rec_reset_tenor"],
            h["swap_pay_fixed_or_floating"],
            _safe_float_val(h["swap_fixed_pay_rt"]),
            h["swap_float_pay_index"], _safe_float_val(h["swap_float_pay_spread"]),
            h["swap_float_pay_reset_tenor"],
            h["swap_other_rec_desc"], h["swap_other_pay_desc"],
            h["swap_termination_dt"],
            _safe_float_val(h["swap_upfront_pmnt"]), _safe_float_val(h["swap_upfront_rcpt"]),
            h["opt_put_or_call"], h["opt_written_or_pur"],
            _safe_float_val(h["opt_share_no"]),
            _safe_float_val(h["opt_exercise_price"]), h["opt_exercise_price_cur"],
            h["opt_exp_dt"], _safe_float_val(h["opt_delta"]),
            h["opt_nested_deriv_type"], h["opt_nested_deriv_title"],
            h["opt_nested_deriv_name"], h["opt_nested_deriv_lei"],
            h["opt_nested_deriv_cusip"], h["opt_nested_deriv_isin"],
            h["opt_nested_deriv_other_id"], h["opt_nested_deriv_other_desc"],
            _safe_float_val(h["opt_nested_deriv_balance"]), h["opt_nested_deriv_units"],
            h["opt_nested_deriv_currency"], h["opt_nested_deriv_exchange_rt"],
            _safe_float_val(h["opt_nested_deriv_val_usd"]), _safe_float_val(h["opt_nested_deriv_pct_val"]),
            h["opt_nested_deriv_asset_cat"], h["opt_nested_deriv_issuer_cat"],
            h["opt_nested_deriv_issuer_desc"], h["opt_nested_deriv_inv_country"],
            h["opt_nested_deriv_cat"], h["opt_nested_deriv_counterparty"],
            h["opt_nested_deriv_counterparty_lei"],
            h["opt_nested_deriv_settlement_dt"],
            _safe_float_val(h["opt_nested_deriv_fwd_amt_sold"]), h["opt_nested_deriv_fwd_cur_sold"],
            _safe_float_val(h["opt_nested_deriv_fwd_amt_pur"]), h["opt_nested_deriv_fwd_cur_pur"],
            h["opt_nested_deriv_swap_rec_fixed_or_float"],
            h["opt_nested_deriv_swap_rec_index"], h["opt_nested_deriv_swap_rec_rt"],
            h["opt_nested_deriv_swap_rec_reset_tenor"],
            h["opt_nested_deriv_swap_pay_fixed_or_float"],
            h["opt_nested_deriv_swap_pay_index"], h["opt_nested_deriv_swap_pay_rt"],
            h["opt_nested_deriv_swap_pay_reset_tenor"],
            h["opt_nested_deriv_swap_termination_dt"],
            _safe_float_val(h["opt_nested_deriv_swap_notional"]), h["opt_nested_deriv_swap_notional_cur"],
            h["deriv_ref_index_name"], h["deriv_ref_index_id"], h["deriv_ref_narrative"],
            h["deriv_ref_issuer_name"], h["deriv_ref_issue_title"],
            h["deriv_ref_cusip"], h["deriv_ref_isin"],
            h["deriv_ref_other_id"], h["deriv_ref_other_desc"],
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)

    _auto_width(ws)


def write_debt_sheet(wb, holdings, styles):
    """Sheet 3: Debt securities with full details."""
    debts = [h for h in holdings if h["has_debt"] == "Y"]
    if not debts:
        return

    ws = wb.create_sheet("Debt Securities")
    headers = [
        "Name", "Title", "CUSIP", "ISIN", "LEI",
        "Balance", "Units", "Currency",
        "Value (USD)", "% of NAV",
        "Maturity Date", "Coupon Kind", "Annualized Rate (%)",
        "Is Default", "Interest In Arrears", "Is Paid In Kind",
        "Is Mandatory Convertible", "Is Contingent Convertible",
        "Ref Instrument Name", "Ref Instrument Title", "Ref CUSIP", "Ref ISIN", "Ref Currency",
        "Conversion Ratio", "Conversion Currency", "Delta",
        "Reset Tenors",
        "Asset Category", "Issuer Category", "Country",
        "Fair Value Level",
    ]
    _write_header_row(ws, headers, styles)

    for row_idx, h in enumerate(debts, 2):
        vals = [
            h["name"], h["title"], h["cusip"], h["isin"], h["lei"],
            _safe_float_val(h["balance"]), h["units"], h["currency"],
            _safe_float_val(h["value_usd"]), _safe_float_val(h["pct_val"]),
            _to_date(h["maturity_dt"]), h["coupon_kind"], _safe_float_val(h["annualized_rt"]),
            h["is_default"], h["are_intrst_pmnts_in_arrs"], h["is_paid_kind"],
            h["is_mandatory_convrtbl"], h["is_contngt_convrtbl"],
            h["debt_ref_name"], h["debt_ref_title"], h["debt_ref_cusip"],
            h["debt_ref_isin"], h["debt_ref_currency"],
            _safe_float_val(h["conv_ratio"]), h["conv_currency"], _safe_float_val(h["debt_delta"]),
            h["reset_tenors"],
            ASSET_CATEGORY_MAP.get(h["asset_cat"], h["asset_cat"]),
            ISSUER_CATEGORY_MAP.get(h["issuer_cat"], h["issuer_cat"]),
            h["inv_country"],
            h["fair_val_level"],
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
            if col_idx == 9 and isinstance(val, (int, float)):
                cell.number_format = styles["money_fmt"]
            elif col_idx == 10 and isinstance(val, (int, float)):
                cell.number_format = styles["pct_fmt"]
            elif col_idx == 13 and isinstance(val, (int, float)):
                cell.number_format = styles["rate_fmt"]

    _auto_width(ws)


def write_repo_sheet(wb, holdings, styles):
    """Sheet 4: Repurchase & reverse repurchase agreements."""
    repos = [h for h in holdings if h["has_repo"] == "Y"]
    if not repos:
        return

    ws = wb.create_sheet("Repo Agreements")
    headers = [
        "Name", "Title", "CUSIP",
        "Transaction Type", "Is Cleared",
        "Counterparty", "Counterparty LEI",
        "Is Tri-Party",
        "Repo Rate (%)", "Maturity Date",
        "Value (USD)", "% of NAV",
        "Collateral Principal", "Collateral Principal Cur",
        "Collateral Value", "Collateral Value Cur",
        "Collateral Category", "Collateral Category Desc",
        "Asset Category", "Country",
    ]
    _write_header_row(ws, headers, styles)

    for row_idx, h in enumerate(repos, 2):
        vals = [
            h["name"], h["title"], h["cusip"],
            h["repo_trans_cat"], h["repo_is_cleared"],
            h["repo_counterparty_name"], h["repo_counterparty_lei"],
            h["repo_is_tri_party"],
            _safe_float_val(h["repo_rate"]), h["repo_maturity_dt"],
            _safe_float_val(h["value_usd"]), _safe_float_val(h["pct_val"]),
            h["repo_collateral_principal"], h["repo_collateral_principal_cur"],
            h["repo_collateral_value"], h["repo_collateral_value_cur"],
            h["repo_collateral_category"], h["repo_collateral_category_desc"],
            ASSET_CATEGORY_MAP.get(h["asset_cat"], h["asset_cat"]),
            h["inv_country"],
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
            if col_idx in (11,) and isinstance(val, (int, float)):
                cell.number_format = styles["money_fmt"]
            elif col_idx == 12 and isinstance(val, (int, float)):
                cell.number_format = styles["pct_fmt"]
            elif col_idx == 9 and isinstance(val, (int, float)):
                cell.number_format = styles["rate_fmt"]

    _auto_width(ws)


def write_lending_sheet(wb, holdings, styles):
    """Sheet 5: Securities lending activity."""
    lending = [h for h in holdings if h["sl_is_loan_by_fund"] == "Y" or h["sl_is_cash_collateral"] == "Y"]
    if not lending:
        return

    ws = wb.create_sheet("Securities Lending")
    headers = [
        "Name", "Title", "CUSIP", "ISIN",
        "Value (USD)",
        "Is Cash Collateral", "Cash Collateral Value",
        "Is Non-Cash Collateral",
        "Is Loan By Fund", "Loan Value",
    ]
    _write_header_row(ws, headers, styles)

    for row_idx, h in enumerate(lending, 2):
        vals = [
            h["name"], h["title"], h["cusip"], h["isin"],
            _safe_float_val(h["value_usd"]),
            h["sl_is_cash_collateral"], _safe_float_val(h["sl_cash_collateral_val"]),
            h["sl_is_non_cash_collateral"],
            h["sl_is_loan_by_fund"], _safe_float_val(h["sl_loan_val"]),
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
            if col_idx in (5, 7, 10) and isinstance(val, (int, float)):
                cell.number_format = styles["money_fmt"]

    _auto_width(ws)


def write_summary_sheet(wb, gen_info, fund_info, filing_info, holdings, ticker, styles, sig_info):
    """Sheet 6: Comprehensive summary."""
    ws = wb.create_sheet("Summary")
    row = 1

    def _section(title):
        nonlocal row
        if row > 1:
            row += 1
        ws.cell(row=row, column=1, value=title).font = styles["section_font"]
        row += 1

    def _kv(label, value, fmt=None):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = styles["bold_font"]
        cell = ws.cell(row=row, column=2, value=value)
        if fmt:
            cell.number_format = fmt
        row += 1

    # Registration
    _section("REGISTRATION INFORMATION")
    _kv("Registrant Name", gen_info.get("reg_name", ""))
    _kv("CIK", gen_info.get("reg_cik", ""))
    _kv("LEI", gen_info.get("reg_lei", ""))
    _kv("File Number", gen_info.get("reg_file_number", ""))
    addr_parts = [gen_info.get("reg_street1", ""), gen_info.get("reg_street2", ""),
                  gen_info.get("reg_city", ""),
                  gen_info.get("reg_state", ""), gen_info.get("reg_zip", ""),
                  gen_info.get("reg_country", "")]
    _kv("Address", ", ".join(p for p in addr_parts if p))
    _kv("Phone", gen_info.get("reg_phone", ""))

    # Fund
    _section("FUND INFORMATION")
    _kv("Series Name", gen_info.get("series_name", ""))
    _kv("Series ID", gen_info.get("series_id", ""))
    _kv("Series LEI", gen_info.get("series_lei", ""))
    _kv("Ticker", ticker)
    _kv("Is Final Filing", gen_info.get("is_final_filing", ""))

    # Reporting period
    _section("REPORTING PERIOD")
    _kv("Reporting Date", gen_info.get("rep_pd_date", ""))
    _kv("Fiscal Year End", gen_info.get("rep_pd_end", ""))
    _kv("Filing Date", filing_info.get("file_date", ""))
    _kv("Accession Number", filing_info.get("accession", ""))
    _kv("Source", "SEC EDGAR NPORT-P")

    # Signature
    if sig_info:
        _section("SIGNATURE")
        _kv("Date Signed", sig_info.get("date_signed", ""))
        _kv("Applicant", sig_info.get("name_of_applicant", ""))
        _kv("Signer", sig_info.get("signer_name", ""))
        _kv("Title", sig_info.get("signer_title", ""))

    # Financials
    fin = fund_info.get("financials", {})
    _section("FINANCIAL SUMMARY")
    _kv("Total Assets", fin.get("tot_assets"), styles["money_fmt"])
    _kv("Total Liabilities", fin.get("tot_liabs"), styles["money_fmt"])
    _kv("Net Assets", fin.get("net_assets"), styles["money_fmt"])
    _kv("Assets Attr. Misc. Securities", fin.get("assets_attr_misc_sec"), styles["money_fmt"])
    _kv("Assets Invested (Sec Lending)", fin.get("assets_invested"), styles["money_fmt"])

    # Borrowing
    bw = fund_info.get("borrowing_within_1yr", {})
    _section("BORROWING (Within 1 Year)")
    _kv("Banks", bw.get("banks"), styles["money_fmt"])
    _kv("Controlled Companies", bw.get("controlled_companies"), styles["money_fmt"])
    _kv("Other Affiliates", bw.get("other_affiliates"), styles["money_fmt"])
    _kv("Other", bw.get("other"), styles["money_fmt"])

    ba = fund_info.get("borrowing_after_1yr", {})
    _section("BORROWING (After 1 Year)")
    _kv("Banks", ba.get("banks"), styles["money_fmt"])
    _kv("Controlled Companies", ba.get("controlled_companies"), styles["money_fmt"])
    _kv("Other Affiliates", ba.get("other_affiliates"), styles["money_fmt"])
    _kv("Other", ba.get("other"), styles["money_fmt"])

    # Other
    of = fund_info.get("other_financials", {})
    _section("OTHER FINANCIAL")
    _kv("Delayed Delivery", of.get("delay_deliv"), styles["money_fmt"])
    _kv("Standby Commitment", of.get("stand_by_commit"), styles["money_fmt"])
    _kv("Liquidation Preference", of.get("liquid_pref"), styles["money_fmt"])
    _kv("Cash Not in C or D", of.get("cash_not_reported"), styles["money_fmt"])
    _kv("Non-Cash Collateral (Fund)", fund_info.get("is_non_cash_collateral", ""))

    # Portfolio reconciliation
    net_assets = fin.get("net_assets") or 0
    total_val = sum(_float(h["value_usd"]) or 0 for h in holdings)
    total_pct = sum(_float(h["pct_val"]) or 0 for h in holdings)
    _section("PORTFOLIO RECONCILIATION")
    _kv("Sum of Holdings (Value USD)", total_val, styles["money_fmt"])
    pct_str = f"{total_val / net_assets * 100:.2f}%" if net_assets else "N/A"
    _kv("Holdings as % of Net Assets", pct_str)
    _kv("Sum of % of NAV", f"{total_pct:.4f}")
    _kv("Total Holdings", len(holdings), styles["int_fmt"])
    _kv("Long Positions", sum(1 for h in holdings if h["payoff_profile"] == "Long"))
    _kv("Short Positions", sum(1 for h in holdings if h["payoff_profile"] == "Short"))
    _kv("N/A Positions (Derivatives)", sum(1 for h in holdings if h["payoff_profile"] == "N/A"))
    _kv("Derivative Positions", sum(1 for h in holdings if h["has_deriv"] == "Y"))
    _kv("  - Forwards", sum(1 for h in holdings if h["deriv_type"] == "FWD"))
    _kv("  - Futures", sum(1 for h in holdings if h["deriv_type"] == "FUT"))
    _kv("  - Swaps", sum(1 for h in holdings if h["deriv_type"] == "SWP"))
    _kv("  - Options/Swaptions/Warrants", sum(1 for h in holdings if h["deriv_type"] in ("OPT", "SWN", "WAR")))
    _kv("Debt Positions", sum(1 for h in holdings if h["has_debt"] == "Y"))
    _kv("Repo/Reverse Repo", sum(1 for h in holdings if h["has_repo"] == "Y"))
    loans = sum(1 for h in holdings if h["sl_is_loan_by_fund"] == "Y")
    _kv("Securities on Loan", loans)

    # Asset category breakdown
    _section("ASSET CATEGORY BREAKDOWN")
    cat_counts = Counter()
    for h in holdings:
        cat_counts[h["asset_cat"] or "(Uncategorized)"] += 1
    for cat, count in cat_counts.most_common():
        _kv(f"  {ASSET_CATEGORY_MAP.get(cat, cat)}", count)

    # VaR
    var = fund_info.get("var_info", {})
    if var.get("designated_index_name"):
        _section("VAR INFORMATION")
        _kv("Designated Index", var.get("designated_index_name", ""))
        _kv("Index Identifier", var.get("index_identifier", ""))

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 60


def write_monthly_returns_sheet(wb, fund_info, styles):
    """Sheet 7: Monthly total returns per class."""
    returns = fund_info.get("monthly_total_returns", [])
    if not returns:
        return

    ws = wb.create_sheet("Monthly Returns")
    headers = ["Class ID", "Month 1 Return (%)", "Month 2 Return (%)", "Month 3 Return (%)"]
    _write_header_row(ws, headers, styles)

    for row_idx, r in enumerate(returns, 2):
        ws.cell(row=row_idx, column=1, value=r["class_id"])
        for col, key in [(2, "return_month1"), (3, "return_month2"), (4, "return_month3")]:
            cell = ws.cell(row=row_idx, column=col, value=r.get(key))
            cell.number_format = styles["rate_fmt"]

    # Other monthly returns (non-derivative)
    oth = fund_info.get("other_monthly_returns", {})
    if oth:
        row = len(returns) + 3
        ws.cell(row=row, column=1, value="Other (Non-Derivative) Returns").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Month")
        ws.cell(row=row, column=2, value="Net Realized Gain")
        ws.cell(row=row, column=3, value="Net Unrealized Appr")
        row += 1
        for key, label in [("month1", "Month 1"), ("month2", "Month 2"), ("month3", "Month 3")]:
            if key in oth:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=oth[key].get("realized")).number_format = styles["money_fmt"]
                ws.cell(row=row, column=3, value=oth[key].get("unrealized")).number_format = styles["money_fmt"]
                row += 1

    _auto_width(ws)


def write_derivative_returns_sheet(wb, fund_info, styles):
    """Sheet 8: Derivative returns by contract type and instrument."""
    dr = fund_info.get("derivative_returns", {})
    if not dr:
        return

    ws = wb.create_sheet("Derivative Returns")
    headers = [
        "Contract Type", "Instrument Type",
        "Mon 1 Realized Gain", "Mon 1 Unrealized Appr",
        "Mon 2 Realized Gain", "Mon 2 Unrealized Appr",
        "Mon 3 Realized Gain", "Mon 3 Unrealized Appr",
    ]
    _write_header_row(ws, headers, styles)

    row = 2
    for contract_type, data in dr.items():
        ws.cell(row=row, column=1, value=contract_type).font = styles["bold_font"]
        ws.cell(row=row, column=2, value="(Aggregate)")
        for col, key in [(3, "mon1_realized"), (4, "mon1_unrealized"),
                         (5, "mon2_realized"), (6, "mon2_unrealized"),
                         (7, "mon3_realized"), (8, "mon3_unrealized")]:
            cell = ws.cell(row=row, column=col, value=data.get(key))
            cell.number_format = styles["money_fmt"]
        row += 1

        for instr_name, instr_data in data.get("instruments", {}).items():
            ws.cell(row=row, column=1, value="")
            ws.cell(row=row, column=2, value=instr_name)
            for col, key in [(3, "mon1_realized"), (4, "mon1_unrealized"),
                             (5, "mon2_realized"), (6, "mon2_unrealized"),
                             (7, "mon3_realized"), (8, "mon3_unrealized")]:
                cell = ws.cell(row=row, column=col, value=instr_data.get(key))
                cell.number_format = styles["money_fmt"]
            row += 1

    _auto_width(ws)


def write_flow_sheet(wb, fund_info, styles):
    """Sheet 9: Monthly flow information."""
    flow = fund_info.get("flow_info", {})
    if not flow:
        return

    ws = wb.create_sheet("Flow Information")
    headers = ["Month", "Sales", "Redemptions", "Reinvestments", "Net Flow"]
    _write_header_row(ws, headers, styles)

    for row_idx, (key, label) in enumerate([("month1", "Month 1"), ("month2", "Month 2"), ("month3", "Month 3")], 2):
        data = flow.get(key, {})
        sales = data.get("sales") or 0
        redemption = data.get("redemption") or 0
        reinvestment = data.get("reinvestment") or 0
        net = sales - redemption + reinvestment
        ws.cell(row=row_idx, column=1, value=label)
        for col, val in [(2, sales), (3, redemption), (4, reinvestment), (5, net)]:
            ws.cell(row=row_idx, column=col, value=val).number_format = styles["money_fmt"]

    _auto_width(ws)


def write_interest_rate_risk_sheet(wb, fund_info, styles):
    """Sheet 10: Interest rate risk by currency."""
    metrics = fund_info.get("currency_metrics", [])
    if not metrics:
        return

    ws = wb.create_sheet("Interest Rate Risk")
    headers = [
        "Currency",
        "DV01 3-Month", "DV01 1-Year", "DV01 5-Year", "DV01 10-Year", "DV01 30-Year",
        "DV100 3-Month", "DV100 1-Year", "DV100 5-Year", "DV100 10-Year", "DV100 30-Year",
    ]
    _write_header_row(ws, headers, styles)

    for row_idx, m in enumerate(metrics, 2):
        ws.cell(row=row_idx, column=1, value=m["currency"])
        for col, key in [(2, "dv01_3m"), (3, "dv01_1y"), (4, "dv01_5y"), (5, "dv01_10y"), (6, "dv01_30y"),
                         (7, "dv100_3m"), (8, "dv100_1y"), (9, "dv100_5y"), (10, "dv100_10y"), (11, "dv100_30y")]:
            cell = ws.cell(row=row_idx, column=col, value=m.get(key))
            cell.number_format = styles["num_fmt"]

    _auto_width(ws)


def write_credit_spread_risk_sheet(wb, fund_info, styles):
    """Sheet 11: Credit spread risk."""
    csr = fund_info.get("credit_spread_risk", {})
    if not any(v is not None for v in csr.values()):
        return

    ws = wb.create_sheet("Credit Spread Risk")
    headers = ["Risk Category", "3-Month", "1-Year", "5-Year", "10-Year", "30-Year"]
    _write_header_row(ws, headers, styles)

    ws.cell(row=2, column=1, value="Investment Grade")
    for col, key in [(2, "inv_grade_3m"), (3, "inv_grade_1y"), (4, "inv_grade_5y"), (5, "inv_grade_10y"), (6, "inv_grade_30y")]:
        ws.cell(row=2, column=col, value=csr.get(key)).number_format = styles["num_fmt"]

    ws.cell(row=3, column=1, value="Non-Investment Grade")
    for col, key in [(2, "non_inv_grade_3m"), (3, "non_inv_grade_1y"), (4, "non_inv_grade_5y"), (5, "non_inv_grade_10y"), (6, "non_inv_grade_30y")]:
        ws.cell(row=3, column=col, value=csr.get(key)).number_format = styles["num_fmt"]

    _auto_width(ws)


def write_borrowers_sheet(wb, fund_info, styles):
    """Sheet 12: Securities lending borrowers."""
    borrowers = fund_info.get("borrowers", [])
    if not borrowers:
        return

    ws = wb.create_sheet("Borrowers")
    headers = ["Name", "LEI", "Aggregate Value"]
    _write_header_row(ws, headers, styles)

    for row_idx, b in enumerate(borrowers, 2):
        ws.cell(row=row_idx, column=1, value=b["name"])
        ws.cell(row=row_idx, column=2, value=b["lei"])
        ws.cell(row=row_idx, column=3, value=b["aggregate_value"]).number_format = styles["money_fmt"]

    _auto_width(ws)


def write_by_asset_category_sheet(wb, holdings, fund_info, styles):
    """Sheet 13: Breakdown by asset category."""
    ws = wb.create_sheet("By Asset Category")
    headers = ["Code", "Category", "# Holdings", "Value (USD)", "% of NAV", "Avg Value"]
    _write_header_row(ws, headers, styles)

    net_assets = (fund_info.get("financials", {}).get("net_assets") or 0)
    cats = Counter()
    cat_vals = defaultdict(float)
    for h in holdings:
        cat = h["asset_cat"] or "(Uncategorized)"
        cats[cat] += 1
        cat_vals[cat] += _float(h["value_usd"]) or 0

    for row_idx, (cat, count) in enumerate(cats.most_common(), 2):
        val = cat_vals[cat]
        ws.cell(row=row_idx, column=1, value=cat)
        ws.cell(row=row_idx, column=2, value=ASSET_CATEGORY_MAP.get(cat, cat))
        ws.cell(row=row_idx, column=3, value=count)
        ws.cell(row=row_idx, column=4, value=val).number_format = styles["money_fmt"]
        ws.cell(row=row_idx, column=5, value=val / net_assets * 100 if net_assets else None).number_format = "0.00"
        ws.cell(row=row_idx, column=6, value=val / count if count else None).number_format = styles["money_fmt"]

    _auto_width(ws)


def write_by_issuer_category_sheet(wb, holdings, fund_info, styles):
    """Sheet 14: Breakdown by issuer category."""
    ws = wb.create_sheet("By Issuer Category")
    headers = ["Code", "Category", "# Holdings", "Value (USD)", "% of NAV"]
    _write_header_row(ws, headers, styles)

    net_assets = (fund_info.get("financials", {}).get("net_assets") or 0)
    cats = Counter()
    cat_vals = defaultdict(float)
    for h in holdings:
        cat = h["issuer_cat"] or "(Uncategorized)"
        cats[cat] += 1
        cat_vals[cat] += _float(h["value_usd"]) or 0

    for row_idx, (cat, count) in enumerate(cats.most_common(), 2):
        val = cat_vals[cat]
        ws.cell(row=row_idx, column=1, value=cat)
        ws.cell(row=row_idx, column=2, value=ISSUER_CATEGORY_MAP.get(cat, cat))
        ws.cell(row=row_idx, column=3, value=count)
        ws.cell(row=row_idx, column=4, value=val).number_format = styles["money_fmt"]
        ws.cell(row=row_idx, column=5, value=val / net_assets * 100 if net_assets else None).number_format = "0.00"

    _auto_width(ws)


def write_by_country_sheet(wb, holdings, fund_info, styles):
    """Sheet 15: Breakdown by country."""
    ws = wb.create_sheet("By Country")
    headers = ["Country", "# Holdings", "Value (USD)", "% of NAV"]
    _write_header_row(ws, headers, styles)

    net_assets = (fund_info.get("financials", {}).get("net_assets") or 0)
    cats = Counter()
    cat_vals = defaultdict(float)
    for h in holdings:
        cat = h["inv_country"] or "(Unknown)"
        cats[cat] += 1
        cat_vals[cat] += _float(h["value_usd"]) or 0

    for row_idx, (cat, count) in enumerate(cats.most_common(), 2):
        val = cat_vals[cat]
        ws.cell(row=row_idx, column=1, value=cat)
        ws.cell(row=row_idx, column=2, value=count)
        ws.cell(row=row_idx, column=3, value=val).number_format = styles["money_fmt"]
        ws.cell(row=row_idx, column=4, value=val / net_assets * 100 if net_assets else None).number_format = "0.00"

    _auto_width(ws)


def write_by_currency_sheet(wb, holdings, fund_info, styles):
    """Sheet 16: Breakdown by currency."""
    ws = wb.create_sheet("By Currency")
    headers = ["Currency", "# Holdings", "Value (USD)", "% of NAV"]
    _write_header_row(ws, headers, styles)

    net_assets = (fund_info.get("financials", {}).get("net_assets") or 0)
    cats = Counter()
    cat_vals = defaultdict(float)
    for h in holdings:
        cat = h["currency"] or "(Unknown)"
        cats[cat] += 1
        cat_vals[cat] += _float(h["value_usd"]) or 0

    for row_idx, (cat, count) in enumerate(cats.most_common(), 2):
        val = cat_vals[cat]
        ws.cell(row=row_idx, column=1, value=cat)
        ws.cell(row=row_idx, column=2, value=count)
        ws.cell(row=row_idx, column=3, value=val).number_format = styles["money_fmt"]
        ws.cell(row=row_idx, column=4, value=val / net_assets * 100 if net_assets else None).number_format = "0.00"

    _auto_width(ws)


def write_top_holdings_sheet(wb, holdings, fund_info, styles):
    """Sheet 17: Top 50 holdings by absolute value."""
    ws = wb.create_sheet("Top 50 Holdings")
    headers = ["Rank", "Name", "Title", "CUSIP", "ISIN", "Asset Category", "Country",
               "Value (USD)", "% of NAV", "Cumulative %"]
    _write_header_row(ws, headers, styles)

    net_assets = (fund_info.get("financials", {}).get("net_assets") or 0)
    sorted_h = sorted(holdings, key=lambda x: abs(_float(x["value_usd"]) or 0), reverse=True)[:50]

    cum_pct = 0.0
    for row_idx, h in enumerate(sorted_h, 2):
        rank = row_idx - 1
        val = _float(h["value_usd"]) or 0
        pct = val / net_assets * 100 if net_assets else 0
        cum_pct += pct
        ws.cell(row=row_idx, column=1, value=rank)
        ws.cell(row=row_idx, column=2, value=h["name"])
        ws.cell(row=row_idx, column=3, value=h["title"])
        ws.cell(row=row_idx, column=4, value=h["cusip"])
        ws.cell(row=row_idx, column=5, value=h["isin"])
        ws.cell(row=row_idx, column=6, value=ASSET_CATEGORY_MAP.get(h["asset_cat"], h["asset_cat"]))
        ws.cell(row=row_idx, column=7, value=h["inv_country"])
        ws.cell(row=row_idx, column=8, value=val).number_format = styles["money_fmt"]
        ws.cell(row=row_idx, column=9, value=pct).number_format = "0.0000"
        ws.cell(row=row_idx, column=10, value=cum_pct).number_format = "0.0000"

    _auto_width(ws)


def write_explanatory_notes_sheet(wb, notes, styles):
    """Sheet 18: Explanatory notes."""
    if not notes:
        return

    ws = wb.create_sheet("Explanatory Notes")
    headers = ["Item", "Note"]
    _write_header_row(ws, headers, styles)

    for row_idx, n in enumerate(notes, 2):
        ws.cell(row=row_idx, column=1, value=n["item"])
        cell = ws.cell(row=row_idx, column=2, value=n["note"])
        cell.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 100


def write_raw_data_sheet(wb, holdings, styles):
    """Sheet 19: FLAT denormalized dump of EVERY field for EVERY holding.
    This is the key sheet for downstream BBG / LLM consumption.
    """
    if not holdings:
        return

    ws = wb.create_sheet("Raw Data (All Fields)")

    # Use all keys from the first holding as headers (exclude internal keys)
    all_keys = [k for k in holdings[0].keys() if not k.startswith("_")]
    # Make human-readable headers
    headers = [k.replace("_", " ").title() for k in all_keys]
    _write_header_row(ws, headers, styles)

    for row_idx, h in enumerate(holdings, 2):
        for col_idx, key in enumerate(all_keys, 1):
            val = h.get(key, "")
            if val == "":
                val = None
            elif isinstance(val, str):
                fval = _safe_float_val(val)
                if isinstance(fval, float):
                    val = fval
            ws.cell(row=row_idx, column=col_idx, value=val)

    _auto_width(ws)


def write_maturity_profile_sheet(wb, holdings, fund_info, styles):
    """Sheet 20: Maturity bucketing for fixed income holdings."""
    debts = [h for h in holdings if h["has_debt"] == "Y" and h["maturity_dt"]]
    if not debts:
        return

    ws = wb.create_sheet("Maturity Profile")
    headers = ["Maturity Bucket", "# Holdings", "Value (USD)", "% of NAV", "Avg Coupon (%)"]
    _write_header_row(ws, headers, styles)

    net_assets = (fund_info.get("financials", {}).get("net_assets") or 0)

    # Parse maturity dates and bucket
    buckets = {
        "0-1 Year": (0, 365),
        "1-3 Years": (366, 1095),
        "3-5 Years": (1096, 1825),
        "5-7 Years": (1826, 2555),
        "7-10 Years": (2556, 3650),
        "10-20 Years": (3651, 7300),
        "20-30 Years": (7301, 10950),
        "30+ Years": (10951, 999999),
        "Undated/Other": (-1, -1),
    }
    bucket_data = {k: {"count": 0, "value": 0.0, "coupons": []} for k in buckets}

    today = date.today()
    for h in debts:
        try:
            mat = datetime.strptime(h["maturity_dt"], "%Y-%m-%d").date()
            days = (mat - today).days
        except (ValueError, TypeError):
            days = -1

        placed = False
        for bname, (lo, hi) in buckets.items():
            if bname == "Undated/Other":
                continue
            if lo <= days <= hi:
                bucket_data[bname]["count"] += 1
                bucket_data[bname]["value"] += _float(h["value_usd"]) or 0
                cpn = _float(h["annualized_rt"])
                if cpn is not None:
                    bucket_data[bname]["coupons"].append(cpn)
                placed = True
                break
        if not placed:
            bucket_data["Undated/Other"]["count"] += 1
            bucket_data["Undated/Other"]["value"] += _float(h["value_usd"]) or 0

    row = 2
    for bname in buckets:
        bd = bucket_data[bname]
        if bd["count"] == 0:
            continue
        ws.cell(row=row, column=1, value=bname)
        ws.cell(row=row, column=2, value=bd["count"])
        ws.cell(row=row, column=3, value=bd["value"]).number_format = styles["money_fmt"]
        ws.cell(row=row, column=4, value=bd["value"] / net_assets * 100 if net_assets else None).number_format = "0.00"
        avg_cpn = sum(bd["coupons"]) / len(bd["coupons"]) if bd["coupons"] else None
        ws.cell(row=row, column=5, value=avg_cpn).number_format = "0.00" if avg_cpn else "General"
        row += 1

    # Total row
    row += 1
    total_count = sum(bd["count"] for bd in bucket_data.values())
    total_val = sum(bd["value"] for bd in bucket_data.values())
    ws.cell(row=row, column=1, value="TOTAL").font = styles["bold_font"]
    ws.cell(row=row, column=2, value=total_count).font = styles["bold_font"]
    ws.cell(row=row, column=3, value=total_val).number_format = styles["money_fmt"]

    _auto_width(ws)


def write_counterparty_exposure_sheet(wb, holdings, styles):
    """Sheet 21: Aggregate derivative exposure by counterparty."""
    derivs = [h for h in holdings if h["has_deriv"] == "Y" and h["counterparty_name"]]
    if not derivs:
        return

    ws = wb.create_sheet("Counterparty Exposure")
    headers = [
        "Counterparty", "LEI",
        "# Positions", "Gross Notional (USD)", "Net Market Value (USD)",
        "Unrealized P&L",
        "Forwards", "Futures", "Swaps", "Options",
    ]
    _write_header_row(ws, headers, styles)

    # Aggregate by counterparty
    cpty_data = defaultdict(lambda: {
        "lei": "", "count": 0, "notional": 0.0, "mv": 0.0, "unreal": 0.0,
        "fwd": 0, "fut": 0, "swp": 0, "opt": 0,
    })
    for h in derivs:
        name = h["counterparty_name"].split(";")[0].strip()  # Take first if multiple
        lei = h["counterparty_lei"].split(";")[0].strip()
        cd = cpty_data[name]
        cd["lei"] = lei
        cd["count"] += 1
        cd["mv"] += _float(h["value_usd"]) or 0
        cd["unreal"] += _float(h["unrealized_appr"]) or 0
        # Notional
        not_val = _float(h["swap_notional_amt"]) or _float(h["fut_notional_amt"]) or _float(h["fwd_amt_cur_pur"]) or _float(h["opt_share_no"]) or 0
        cd["notional"] += abs(not_val)
        dt = h["deriv_type"]
        if dt == "FWD":
            cd["fwd"] += 1
        elif dt == "FUT":
            cd["fut"] += 1
        elif dt == "SWP":
            cd["swp"] += 1
        elif dt in ("OPT", "SWN", "WAR"):
            cd["opt"] += 1

    # Sort by absolute market value
    sorted_cptys = sorted(cpty_data.items(), key=lambda x: abs(x[1]["mv"]), reverse=True)

    red_font = Font(color="CC0000", size=10)
    for row_idx, (name, cd) in enumerate(sorted_cptys, 2):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=cd["lei"])
        ws.cell(row=row_idx, column=3, value=cd["count"])
        ws.cell(row=row_idx, column=4, value=cd["notional"]).number_format = styles["money_fmt"]
        cell_mv = ws.cell(row=row_idx, column=5, value=cd["mv"])
        cell_mv.number_format = styles["money_fmt"]
        if cd["mv"] < 0:
            cell_mv.font = red_font
        cell_unreal = ws.cell(row=row_idx, column=6, value=cd["unreal"])
        cell_unreal.number_format = styles["money_fmt"]
        if cd["unreal"] < 0:
            cell_unreal.font = red_font
        ws.cell(row=row_idx, column=7, value=cd["fwd"])
        ws.cell(row=row_idx, column=8, value=cd["fut"])
        ws.cell(row=row_idx, column=9, value=cd["swp"])
        ws.cell(row=row_idx, column=10, value=cd["opt"])

    _auto_width(ws)


def write_instrument_type_breakdown_sheet(wb, holdings, fund_info, styles):
    """Sheet 22: Breakdown by unified instrument type classification."""
    ws = wb.create_sheet("By Instrument Type")
    headers = ["Instrument Type", "# Holdings", "Value (USD)", "% of NAV", "Avg Value"]
    _write_header_row(ws, headers, styles)

    net_assets = (fund_info.get("financials", {}).get("net_assets") or 0)
    cats = Counter()
    cat_vals = defaultdict(float)
    for h in holdings:
        cat = h["instrument_type"]
        cats[cat] += 1
        cat_vals[cat] += _float(h["value_usd"]) or 0

    for row_idx, (cat, count) in enumerate(cats.most_common(), 2):
        val = cat_vals[cat]
        ws.cell(row=row_idx, column=1, value=cat)
        ws.cell(row=row_idx, column=2, value=count)
        ws.cell(row=row_idx, column=3, value=val).number_format = styles["money_fmt"]
        ws.cell(row=row_idx, column=4, value=val / net_assets * 100 if net_assets else None).number_format = "0.00"
        ws.cell(row=row_idx, column=5, value=val / count if count else None).number_format = styles["money_fmt"]

    _auto_width(ws)


def write_data_validation_sheet(wb, gen_info, fund_info, holdings, styles):
    """Sheet 23: Data quality checks and reconciliation."""
    ws = wb.create_sheet("Data Validation")
    row = 1

    bold = styles["bold_font"]
    green = Font(bold=True, color="006600", size=10)
    red = Font(bold=True, color="CC0000", size=10)

    def _check(label, value, status, detail=""):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        cell = ws.cell(row=row, column=3, value=status)
        cell.font = green if status == "PASS" else red
        if detail:
            ws.cell(row=row, column=4, value=detail)
        row += 1

    ws.cell(row=row, column=1, value="DATA VALIDATION CHECKS").font = styles["section_font"]
    row += 2

    # Check 1: Holdings count
    _check("Total Holdings Parsed", len(holdings), "INFO")

    # Check 2: NAV reconciliation
    net_assets = fund_info.get("financials", {}).get("net_assets") or 0
    total_val = sum(_float(h["value_usd"]) or 0 for h in holdings)
    if net_assets:
        pct = total_val / net_assets * 100
        status = "PASS" if 90 <= pct <= 110 else "WARN"
        _check("Sum Holdings / NAV", f"{pct:.2f}%", status, f"Holdings={total_val:,.0f} NAV={net_assets:,.0f}")
    else:
        _check("Sum Holdings / NAV", "N/A", "WARN", "Net assets is zero or missing")

    # Check 3: Sum of pctVal
    total_pct = sum(_float(h["pct_val"]) or 0 for h in holdings)
    status = "PASS" if 0.9 <= total_pct <= 1.1 else "WARN"
    _check("Sum of % of NAV", f"{total_pct:.4f}", status, "Expected ~1.0 (or ~100 if in percent)")

    # Check 4: Missing CUSIPs
    no_cusip = sum(1 for h in holdings if not h["cusip"] or h["cusip"] in ("000000000", "N/A"))
    total = len(holdings)
    pct_missing = no_cusip / total * 100 if total else 0
    _check("Missing/Zero CUSIPs", f"{no_cusip} / {total} ({pct_missing:.1f}%)",
           "PASS" if pct_missing < 30 else "WARN",
           "Derivatives typically have 000000000")

    # Check 5: Missing ISINs
    no_isin = sum(1 for h in holdings if not h["isin"])
    pct_missing = no_isin / total * 100 if total else 0
    _check("Missing ISINs", f"{no_isin} / {total} ({pct_missing:.1f}%)",
           "PASS" if pct_missing < 50 else "WARN")

    # Check 6: Defaults
    defaults = [h for h in holdings if h.get("is_default") == "Y"]
    _check("Securities in Default", len(defaults),
           "WARN" if defaults else "PASS",
           "; ".join(h["title"][:40] for h in defaults[:5]))

    # Check 7: Interest payments in arrears
    arrears = [h for h in holdings if h.get("are_intrst_pmnts_in_arrs") == "Y"]
    _check("Interest in Arrears", len(arrears),
           "WARN" if arrears else "PASS",
           "; ".join(h["title"][:40] for h in arrears[:5]))

    # Check 8: Negative values (short positions)
    negatives = sum(1 for h in holdings if (_float(h["value_usd"]) or 0) < 0)
    _check("Negative Value Holdings", negatives, "INFO", "Short positions and written options")

    # Check 9: Fair value level 3
    lvl3 = sum(1 for h in holdings if h["fair_val_level"] == "3")
    _check("Fair Value Level 3 (Illiquid)", lvl3,
           "WARN" if lvl3 > 0 else "PASS",
           "Inputs based on unobservable data")

    # Check 10: Restricted securities
    restricted = sum(1 for h in holdings if h["is_restricted_sec"] == "Y")
    _check("Restricted Securities (144A etc.)", restricted, "INFO")

    # Check 11: Convertible bonds
    conv = sum(1 for h in holdings if h.get("is_mandatory_convrtbl") == "Y" or h.get("is_contngt_convrtbl") == "Y")
    _check("Convertible Bonds", conv, "INFO")

    # Check 12: Securities on loan
    on_loan = sum(1 for h in holdings if h["sl_is_loan_by_fund"] == "Y")
    _check("Securities on Loan", on_loan, "INFO")

    row += 1
    ws.cell(row=row, column=1, value="INSTRUMENT TYPE COUNTS").font = styles["section_font"]
    row += 1
    type_counts = Counter(h["instrument_type"] for h in holdings)
    for itype, count in type_counts.most_common():
        ws.cell(row=row, column=1, value=itype)
        ws.cell(row=row, column=2, value=count)
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60


def _find_unmapped_fields(holdings, ns_str):
    """Compare raw XML against parsed dict for each holding.

    Returns a list of dicts:
      {holding_idx, name, title, cusip, xml_path, attr_or_text, value}
    for every XML value not found in the parsed holding dict.
    """
    unmapped = []
    ns = ns_str  # e.g. "{http://www.sec.gov/edgar/nport}"

    for idx, h in enumerate(holdings):
        raw = h.get("_raw_xml")
        if raw is None:
            continue

        # Build set of all captured values (split compound "; " values)
        captured = set()
        for k, v in h.items():
            if k.startswith("_") or not v or v == "":
                continue
            if isinstance(v, str):
                for part in v.split("; "):
                    p = part.strip()
                    if p:
                        captured.add(p)
                        # Also add sub-tokens for compound reset-tenor strings
                        # e.g. "Month 3 (reset Month 3)" contains "Month", "3"
                        for token in p.replace("(", " ").replace(")", " ").split():
                            captured.add(token)
            else:
                captured.add(str(v))

        def _walk(elem, path_parts):
            tag = elem.tag.replace(ns, "") if ns else elem.tag
            current_path = "/".join(path_parts + [tag])

            # Check text
            if elem.text and elem.text.strip():
                val = elem.text.strip()
                if val not in captured:
                    unmapped.append({
                        "holding_idx": idx + 1,
                        "name": h.get("name", ""),
                        "title": h.get("title", ""),
                        "cusip": h.get("cusip", ""),
                        "xml_path": current_path,
                        "attr_or_text": "text",
                        "value": val,
                    })

            # Check attributes
            for attr_name, attr_val in elem.attrib.items():
                if attr_val and attr_val.strip():
                    val = attr_val.strip()
                    if val not in captured:
                        unmapped.append({
                            "holding_idx": idx + 1,
                            "name": h.get("name", ""),
                            "title": h.get("title", ""),
                            "cusip": h.get("cusip", ""),
                            "xml_path": current_path + f"@{attr_name}",
                            "attr_or_text": "attr",
                            "value": val,
                        })

            # Recurse children
            for child in elem:
                _walk(child, path_parts + [tag])

        _walk(raw, [])

    return unmapped


def write_unmapped_fields_sheet(wb, holdings, ns_str, styles):
    """Sheet 24: Safety-net sheet listing any XML values NOT captured in the
    parsed dict.  If this sheet is empty, we have 100% field coverage for
    the filing.  If not, the user can see exactly which XML paths carry data
    that our parser doesn't yet map.
    """
    unmapped = _find_unmapped_fields(holdings, ns_str)

    ws = wb.create_sheet("Unmapped XML Fields")
    headers = [
        "Holding #", "Name", "Title", "CUSIP",
        "XML Path", "Type", "Value", "Occurrences",
    ]
    _write_header_row(ws, headers, styles)

    if not unmapped:
        ws.cell(row=2, column=1, value="No unmapped fields found — 100% coverage!")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws["A2"].font = Font(bold=True, color="008000", size=12)
        _auto_width(ws)
        return

    # Aggregate: count occurrences per (xml_path, value) to keep sheet concise
    from collections import OrderedDict
    agg = OrderedDict()
    for u in unmapped:
        key = (u["xml_path"], u["value"])
        if key not in agg:
            agg[key] = {
                "first_holding": u["holding_idx"],
                "name": u["name"],
                "title": u["title"],
                "cusip": u["cusip"],
                "xml_path": u["xml_path"],
                "attr_or_text": u["attr_or_text"],
                "value": u["value"],
                "count": 0,
            }
        agg[key]["count"] += 1

    # Sort by count descending then path
    rows = sorted(agg.values(), key=lambda r: (-r["count"], r["xml_path"]))

    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for row_idx, r in enumerate(rows, 2):
        vals = [
            r["first_holding"], r["name"], r["title"], r["cusip"],
            r["xml_path"], r["attr_or_text"], r["value"], r["count"],
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
            if r["count"] >= 10:
                cell.fill = warn_fill

    _auto_width(ws)

    # Summary row at top (insert after header)
    total_values = sum(r["count"] for r in rows)
    unique_paths = len(rows)
    ws.insert_rows(2)
    ws.cell(row=2, column=1,
            value=f"Summary: {total_values} unmapped values across {unique_paths} unique XML paths")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"].font = Font(bold=True, color="CC0000", size=11)


# ============================================================================
# Excel Writing: Orchestrator
# ============================================================================

def write_excel(gen_info, fund_info, holdings, notes, filing_info, ticker, output_path, sig_info=None, ns_str=""):
    """Create the comprehensive Excel workbook."""
    wb, styles = create_workbook()

    # Sheet 1: Holdings (uses the default active sheet)
    if holdings:
        write_holdings_sheet(wb, holdings, styles, ticker)
    else:
        ws = wb.active
        ws.title = f"{ticker} Holdings"
        ws.cell(row=1, column=1, value="No holdings found in this filing.")

    # Sheet 2: Derivatives
    write_derivatives_sheet(wb, holdings, styles)

    # Sheet 3: Debt Securities
    write_debt_sheet(wb, holdings, styles)

    # Sheet 4: Repo / Reverse Repo Agreements
    write_repo_sheet(wb, holdings, styles)

    # Sheet 5: Securities Lending
    write_lending_sheet(wb, holdings, styles)

    # Sheet 6: Summary
    write_summary_sheet(wb, gen_info, fund_info, filing_info, holdings, ticker, styles, sig_info or {})

    # Sheet 7: Monthly Returns
    write_monthly_returns_sheet(wb, fund_info, styles)

    # Sheet 8: Derivative Returns
    write_derivative_returns_sheet(wb, fund_info, styles)

    # Sheet 9: Flow Information
    write_flow_sheet(wb, fund_info, styles)

    # Sheet 10: Interest Rate Risk
    write_interest_rate_risk_sheet(wb, fund_info, styles)

    # Sheet 11: Credit Spread Risk
    write_credit_spread_risk_sheet(wb, fund_info, styles)

    # Sheet 12: Borrowers
    write_borrowers_sheet(wb, fund_info, styles)

    # Sheet 13-16: Breakdowns
    write_by_asset_category_sheet(wb, holdings, fund_info, styles)
    write_by_issuer_category_sheet(wb, holdings, fund_info, styles)
    write_by_country_sheet(wb, holdings, fund_info, styles)
    write_by_currency_sheet(wb, holdings, fund_info, styles)

    # Sheet 17: Top Holdings
    write_top_holdings_sheet(wb, holdings, fund_info, styles)

    # Sheet 18: Explanatory Notes
    write_explanatory_notes_sheet(wb, notes, styles)

    # Sheet 19: Raw Data (flat dump for BBG/LLM)
    write_raw_data_sheet(wb, holdings, styles)

    # Sheet 20: Maturity Profile
    write_maturity_profile_sheet(wb, holdings, fund_info, styles)

    # Sheet 21: Counterparty Exposure
    write_counterparty_exposure_sheet(wb, holdings, styles)

    # Sheet 22: By Instrument Type
    write_instrument_type_breakdown_sheet(wb, holdings, fund_info, styles)

    # Sheet 23: Data Validation
    write_data_validation_sheet(wb, gen_info, fund_info, holdings, styles)

    # Sheet 24: Unmapped XML Fields (safety net)
    if ns_str and holdings and holdings[0].get("_raw_xml") is not None:
        write_unmapped_fields_sheet(wb, holdings, ns_str, styles)

    # Clean up internal keys before saving
    for h in holdings:
        h.pop("_raw_xml", None)

    wb.save(output_path)
    return wb


# ============================================================================
# CLI & Main
# ============================================================================

def build_cli():
    parser = argparse.ArgumentParser(
        description="Extract exhaustive mutual fund holdings from SEC EDGAR N-PORT filings"
    )
    parser.add_argument("ticker", help="Mutual fund ticker symbol (e.g., PBAIX)")
    parser.add_argument("--output", "-o", help="Output Excel file path (default: <TICKER>_nport.xlsx)")
    parser.add_argument("--user-agent", help="Your name and email for SEC fair access policy")
    parser.add_argument("--date", help="Target reporting date (YYYY-MM-DD) for historical lookup")
    parser.add_argument("--all-filings", action="store_true", help="Download all available N-PORT filings")
    return parser


def main():
    parser = build_cli()
    args = parser.parse_args()

    user_agent = args.user_agent or DEFAULT_USER_AGENT
    headers = {"User-Agent": user_agent}

    ticker = args.ticker.upper()

    # Step 1: Look up ticker
    info = lookup_ticker(ticker, headers)
    print(f"Found: CIK={info['cik']}, Series={info['series_id']}, Class={info['class_id']}")

    if args.all_filings:
        filings = find_all_nport_filings(info["cik"], info["series_id"], headers)
        print(f"Found {len(filings)} NPORT-P filings")
        for filing in filings:
            _process_filing(filing, info, ticker, headers, args)
    else:
        filing = find_nport_filing(info["cik"], info["series_id"], headers, target_date=args.date)
        print(f"Filing: {filing['file_date']}, Report: {filing.get('report_date', 'N/A')}")
        _process_filing(filing, info, ticker, headers, args)


def _process_filing(filing, info, ticker, headers, args):
    """Download, parse, and write Excel for a single filing."""
    # Step 2: Download XML
    root = download_nport_xml(filing["cik"], filing["accession"], filing["primary_document"], headers)
    ns = _ns(root)

    # Step 3: Parse all sections
    gen_info = parse_gen_info(root, ns)
    fund_info = parse_fund_info(root, ns)
    holdings = parse_holdings(root, ns)
    notes = parse_explanatory_notes(root, ns)
    sig_info = parse_signature(root, ns)

    deriv_count = sum(1 for h in holdings if h['has_deriv'] == 'Y')
    debt_count = sum(1 for h in holdings if h['has_debt'] == 'Y')
    repo_count = sum(1 for h in holdings if h['has_repo'] == 'Y')
    fwd_count = sum(1 for h in holdings if h['deriv_type'] == 'FWD')
    fut_count = sum(1 for h in holdings if h['deriv_type'] == 'FUT')
    swap_count = sum(1 for h in holdings if h['deriv_type'] == 'SWP')
    opt_count = sum(1 for h in holdings if h['deriv_type'] in ('OPT', 'SWN', 'WAR'))

    print(f"Parsed: {len(holdings)} holdings")
    print(f"  Debt: {debt_count}, Derivatives: {deriv_count} "
          f"(FWD:{fwd_count} FUT:{fut_count} SWP:{swap_count} OPT:{opt_count}), "
          f"Repos: {repo_count}, Notes: {len(notes)}")

    # Step 4: Determine output path
    if args.output and not args.all_filings:
        output_path = args.output
    else:
        report_date = gen_info.get("rep_pd_date") or filing.get("report_date") or filing["file_date"]
        output_path = f"{ticker}_nport_{report_date}.xlsx"

    # Step 5: Write Excel
    write_excel(gen_info, fund_info, holdings, notes, filing, ticker, output_path, sig_info, ns_str=ns)

    # Print summary
    net_assets = (fund_info.get("financials", {}).get("net_assets") or 0)
    print(f"\nSaved to {output_path}")
    print(f"  Fund: {gen_info.get('series_name', 'N/A')}")
    print(f"  As of: {gen_info.get('rep_pd_date', 'N/A')}")
    print(f"  Net Assets: ${net_assets:,.2f}")
    print(f"  Holdings: {len(holdings)}")
    if deriv_count:
        print(f"  Derivatives: {deriv_count} (FWD:{fwd_count} FUT:{fut_count} SWP:{swap_count} OPT:{opt_count})")
    if debt_count:
        print(f"  Debt Securities: {debt_count}")
    if repo_count:
        print(f"  Repo/Reverse Repo: {repo_count}")
    print(f"  Excel Sheets: up to 23")


if __name__ == "__main__":
    main()
